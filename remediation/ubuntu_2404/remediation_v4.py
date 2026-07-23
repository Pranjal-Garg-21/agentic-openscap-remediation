#!/usr/bin/env python3
"""
CIS Benchmark Multi-Model Remediation Pipeline v3.3
=====================================================
New in v3.3 -- multi-pass remediation agent:

  5. MULTI-PASS RETRY WITH ERROR FEEDBACK:
     - Each KEEP rule now gets up to MAX_PASSES attempts (default 3) instead
       of one shot. On script_error / oscap_fail, the exact error is fed
       back to the model verbatim ("here's the script you wrote, here's
       what happened, fix it") instead of just logging a failure and moving
       on. Human approval is only asked once per rule (pass 1); retries
       within the same rule apply automatically since the human already
       signed off on remediating this rule.

  6. SUDOERS CORRUPTION GUARD:
     - The MIN/MAX qwen runs showed one bad script writing a malformed
       /etc/sudoers.d file, which silently broke `sudo` parsing for every
       rule that ran afterward (stderr showed the same syntax error on ~50
       unrelated rules downstream of it). Any script that mentions
       sudoers/visudo now gets a pre-attempt snapshot; after it runs,
       `visudo -c` is checked, and a broken sudoers state is rolled back
       immediately instead of poisoning the rest of the run.

  7. CLARIFYING QUESTIONS:
     - The prompt now tells the model it may respond with a CLARIFY block
       instead of a script when it has a genuine, non-technical doubt (e.g.
       ambiguity about what the profile intends). The pipeline pauses, shows
       the question as multiple-choice with a free-text fallback, and folds
       the answer back into the retry prompt. Capped at one clarification
       per rule so it can't stall indefinitely.

  8. PERMISSION GATE FOR INTERACTIVE SUDO:
     - If a script fails specifically because it needs a password (no
       NOPASSWD rule covers that command), the pipeline no longer just logs
       a script_error -- it asks the human for explicit permission, and only
       if granted, re-runs the script attached to the real terminal so sudo
       can prompt for the password directly.

Fixes vs v3.1:

  1. QUERY CRASH ('choices' KeyError):
     - query_lab_model() no longer accesses data["choices"][0]... before
       checking which key shape the response actually has. The detection
       logic (choices / message / response / unknown) now runs first, inside
       a try/except, so a non-OpenAI-shaped response no longer kills the
       whole rule with [QUERY ERROR] 'choices'.

  2. PERMISSION DENIED ON APPLIED SCRIPTS:
     - Scripts from the LLM are now executed with `sudo bash -c <script>`
       instead of a bare `bash -c <script>`. Previously the subprocess ran
       as whatever unprivileged user launched the pipeline, so anything
       touching /etc, /var/log, apt, aideinit, etc. failed with
       "Permission denied" even though the script itself said `sudo ...`
       for some lines but not others.
     - Recommended: run this whole pipeline with a user that has passwordless
       sudo (NOPASSWD) configured for --auto runs, OR just run the pipeline
       itself with `sudo python3 remediationv3.py ...`. Either works; don't
       do both (harmless but redundant).

  4. --only-min-delta (NEW):
     - When you've already run a full MIN-profile pass (all 51 KEEP-under-MIN
       rules tested per model) and now want to test MAX, you don't need to
       re-test those same 51 rules if MAX's KEEP set is a superset of MIN's
       KEEP set (true for this ground truth file -- verified: every rule
       that's KEEP under MIN is also KEEP under MAX).
     - Pass --only-min-delta together with --profile MAX to restrict the run
       to ONLY the 12 rules where the MIN decision was SKIP (these become
       the new ground to cover for MAX). The other 51 rules are excluded
       from rule_set entirely before the model loop starts -- not run,
       not logged as skipped, just not part of this run at all.
     - Guardrail: using --only-min-delta with --profile MIN is refused,
       since it would filter down to nothing meaningful (every remaining
       rule is SKIP-under-MIN by definition, so nothing would be sent to
       the LLM).

  3. ONLY KEEP-DECISION RULES ARE SENT TO THE LLM:
     - Previously ALL 63 rules were queried regardless of the ground truth
       KEEP/SKIP decision for the selected profile (MAX/MIN), with the
       decision only logged after the fact.
     - Now, before querying the model, each rule's ground-truth decision for
       the CHOSEN profile is checked. If it's SKIP, the rule is recorded as
       "skipped_by_ground_truth" in the results WITHOUT ever calling the LLM
       or touching the system, and the loop moves straight to the next rule.
     - Only rules whose ground-truth decision is KEEP are sent to the model
       for remediation, applied, and oscap-verified.

Usage:
  python3 remediationv3.py --probe                  # diagnose LAB_URL endpoints only
  python3 remediationv3.py                          # interactive, human approval
  python3 remediationv3.py --auto                   # auto-approve all fixes
  python3 remediationv3.py --model qwen2.5:7b       # single model only
  python3 remediationv3.py --profile MAX            # skip the profile prompt
  python3 remediationv3.py --snapshot               # prompt snapshot restore between models
  python3 remediationv3.py --profile MAX --only-min-delta --auto --snapshot
                                                      # only test the 12 rules
                                                      # that were SKIP under MIN
                                                      # (assumes MIN run already done)
  python3 remediationv3.py --profile MIN --retry-failed remediation_results/qwen2_5_7b_MIN_...json --auto
                                                      # only re-run rules that
                                                      # weren't oscap_pass in a
                                                      # prior result file for
                                                      # this profile
  python3 remediationv3.py --profile MIN --only-rules sudo_custom_logfile,aide_build_database --auto
                                                      # only run the named
                                                      # rules (short name or
                                                      # full xccdf id, comma
                                                      # separated)
"""

import os, re, sys, json, time, datetime, subprocess, xml.etree.ElementTree as ET
import requests, urllib3
from openpyxl import load_workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

LAB_URL  = os.environ.get("LAB_URL",  "https://10.1.96.96:8443")
LAB_USER = os.environ.get("LAB_USER", "user")
LAB_PASS = os.environ.get("LAB_PASS", "H72j8n19sna")

MODELS = [
   # "qwen2.5:7b",
  #  "gemma2:latest",
    #"mistral:latest",
    #"granite4.1:8b",
    "gpt-oss:latest",
]

def resolve_path(p):
    if not p or os.path.isabs(p): return p
    if os.path.exists(p): return os.path.abspath(p)
    s_dir = os.path.dirname(os.path.abspath(__file__))
    for base in [s_dir, os.path.join(s_dir, ".."), os.path.join(s_dir, "..", "..")]:
        cand = os.path.abspath(os.path.join(base, p))
        if os.path.exists(cand): return cand
    return os.path.abspath(os.path.join(s_dir, "..", "..", p))

SCAN_RESULT_XML   = "agent-test.xml"
GROUND_TRUTH_XLSX = "CIS_Ground_Truth_FULL_MAX_MIN.xlsx"
GROUND_TRUTH_SHEET_CANDIDATES = ["Ground Truth", "Ground Truth Grid"]
RESULTS_DIR       = "remediation_results"
BENCHMARK_XML     = os.path.expanduser(
    "~/Downloads/scap-security-guide-0.1.76/ssg-ubuntu2404-ds.xml")
SNAPSHOT_NAME     = "baseline-63-rules-broken"

# If True, applied scripts run via `sudo bash -c <script>`.
# Set to False only if you are already running this whole pipeline as root
# (e.g. via `sudo python3 remediationv3.py ...`) -- in that case adding sudo
# again is harmless but unnecessary.
ELEVATE_WITH_SUDO = True

# Max attempts per KEEP rule in the multi-pass agent loop (query -> apply ->
# verify, retrying with error feedback on failure). 1 = old single-pass
# behavior.
MAX_PASSES = 3

# Default per-attempt script timeout, in seconds. Most CIS fixes are a few
# lines of sed/systemctl and finish in well under a second -- but a few
# rules do genuinely slow, filesystem-wide work and need much more room.
# Silently timing out on these (as happened with aide_build_database, which
# scans the whole disk to build its integrity database) burned all 3 passes
# with zero visible feedback before this table existed.
DEFAULT_SCRIPT_TIMEOUT = 120
RULE_TIMEOUTS = {
    "aide_build_database": 3600,  # aideinit walks the whole filesystem -- 900s wasn't enough
}

# Pre-verified, known-correct scripts for rules that kept failing even with
# a text hint -- for these, a description of the bug isn't enough anymore,
# the model needs a concrete correct script to anchor to instead of writing
# one from scratch each pass. Keyed by SHORT rule name. Injected into the
# prompt as "use this exact script" rather than "here's what went wrong."
KNOWN_FIXES = {
    "sudo_require_reauthentication": """\
if dpkg-query --show --showformat='${db:Status-Status}\\n' 'sudo' 2>/dev/null | grep -q '^installed'; then
    echo 'Defaults timestamp_timeout=0' > /etc/sudoers.d/reauthenticate_sudo
    chmod 0440 /etc/sudoers.d/reauthenticate_sudo
fi""",
    "sudo_custom_logfile": """\
if dpkg-query --show --showformat='${db:Status-Status}\\n' 'sudo' 2>/dev/null | grep -q '^installed'; then
    mkdir -p /var/log/sudo
    touch /var/log/sudo.log
    chmod 640 /var/log/sudo.log
    chown root:adm /var/log/sudo.log
    echo 'Defaults logfile="/var/log/sudo.log"' > /etc/sudoers.d/01-cis-sudo-logfile
    chmod 0440 /etc/sudoers.d/01-cis-sudo-logfile
fi""",
    "service_nftables_enabled": """\
if dpkg-query --show --showformat='${db:Status-Status}\\n' 'nftables' 2>/dev/null | grep -q '^installed'; then
    systemctl unmask nftables
    systemctl enable nftables
    systemctl start nftables
fi""",
    "accounts_tmout": """\
cat > /etc/profile.d/tmout.sh << 'EOF'
TMOUT=600
readonly TMOUT
export TMOUT
EOF
chmod 644 /etc/profile.d/tmout.sh""",
    "accounts_passwords_pam_faillock_deny": """\
if grep -qE '^deny[[:space:]]*=' /etc/security/faillock.conf 2>/dev/null; then
    sed -i -E 's/^deny[[:space:]]*=.*/deny = 4/' /etc/security/faillock.conf
else
    echo 'deny = 4' >> /etc/security/faillock.conf
fi""",
    "accounts_passwords_pam_faillock_unlock_time": """\
if grep -qE '^unlock_time[[:space:]]*=' /etc/security/faillock.conf 2>/dev/null; then
    sed -i -E 's/^unlock_time[[:space:]]*=.*/unlock_time = 900/' /etc/security/faillock.conf
else
    echo 'unlock_time = 900' >> /etc/security/faillock.conf
fi""",
    "accounts_passwords_pam_faillock_enabled": """\
conf_name=cac_faillock
if [ ! -f /usr/share/pam-configs/"$conf_name" ]; then
cat << 'EOF' > /usr/share/pam-configs/cac_faillock
Name: Enable pam_faillock to deny access
Default: yes
Priority: 0
Auth-Type: Primary
Auth:
        [default=die] pam_faillock.so authfail
EOF
fi
if [ ! -f /usr/share/pam-configs/"$conf_name"_notify ]; then
cat << 'EOF' > /usr/share/pam-configs/cac_faillock_notify
Name: Notify of failed login attempts and reset count upon success
Default: yes
Priority: 1025
Auth-Type: Primary
Auth:
        requisite pam_faillock.so preauth
Account-Type: Primary
Account:
        required pam_faillock.so
EOF
fi
DEBIAN_FRONTEND=noninteractive pam-auth-update --enable cac_faillock --enable cac_faillock_notify""",
    "accounts_password_pam_minlen": """\
if grep -qE '^minlen[[:space:]]*=' /etc/security/pwquality.conf 2>/dev/null; then
    sed -i -E 's/^minlen[[:space:]]*=.*/minlen = 12/' /etc/security/pwquality.conf
else
    echo 'minlen = 12' >> /etc/security/pwquality.conf
fi""",
    "accounts_password_pam_ucredit": """\
if grep -qE '^ucredit[[:space:]]*=' /etc/security/pwquality.conf 2>/dev/null; then
    sed -i -E 's/^ucredit[[:space:]]*=.*/ucredit = -1/' /etc/security/pwquality.conf
else
    echo 'ucredit = -1' >> /etc/security/pwquality.conf
fi""",
    "accounts_password_pam_dcredit": """\
if grep -qE '^dcredit[[:space:]]*=' /etc/security/pwquality.conf 2>/dev/null; then
    sed -i -E 's/^dcredit[[:space:]]*=.*/dcredit = -1/' /etc/security/pwquality.conf
else
    echo 'dcredit = -1' >> /etc/security/pwquality.conf
fi""",
}

# Per-rule hints for known persistent failures. Generic error feedback alone
# often isn't enough for a 7B model to fix these -- it just reproduces the
# same broken script on every retry (confirmed: aide_build_database returned
# the identical exit code 21 on all 3 passes; several pam_* rules failed
# identically 3x with zero variation attempted). These hints hand the model
# the specific fact it's missing, up front, instead of hoping self-correction
# finds it. Keyed by SHORT rule name (without the xccdf prefix).
RULE_HINTS = {
    "sudo_require_reauthentication":
        "Any line written into /etc/sudoers or /etc/sudoers.d/* MUST start "
        "with the `Defaults` keyword, e.g. `Defaults timestamp_timeout=0`. "
        "IMPORTANT -- the VALUE matters, not just the syntax: "
        "`timestamp_timeout=0` forces sudo to ask for a password every "
        "single time (this is what 'require reauthentication' means). "
        "A NEGATIVE value like -1 means the exact opposite -- per the sudo "
        "manual, a value less than 0 makes the cached credential NEVER "
        "expire, so the user is never asked to reauthenticate again. Do "
        "NOT use -1 for this rule. ALSO: any file you create under "
        "/etc/sudoers.d/ must be mode 0440 (owner and group read-only, no "
        "write, no world access) -- `visudo -c` rejects any sudoers.d file "
        "with different permissions. Always `chmod 0440` the file you "
        "write, in the same script that creates it.",
    "sudo_custom_logfile":
        "Any line written into /etc/sudoers or /etc/sudoers.d/* MUST start "
        "with the `Defaults` keyword, e.g. `Defaults logfile=\"/var/log/sudo.log\"`. "
        "A bare `logfile /var/log/sudo.log` line is invalid syntax and "
        "breaks sudo for the entire system, not just this rule. ALSO: any "
        "file you create under /etc/sudoers.d/ must be mode 0440 (owner "
        "and group read-only, no write, no world access) -- `visudo -c` "
        "rejects any sudoers.d file with different permissions. Always "
        "`chmod 0440` the file you write, in the same script that creates "
        "it -- don't wait to be told this failed, do it up front.",
    "aide_build_database":
        "If `aideinit` returns a non-zero exit code, don't just retry the "
        "same command -- first check for and remove any stale "
        "/var/lib/aide/aide.db or aide.db.new from a previous attempt, and "
        "run `aideinit -y -f` with output captured so the real underlying "
        "error is visible instead of just the exit code.",
    "service_nftables_enabled":
        "This service is likely masked by default (check `systemctl status "
        "nftables`). You must run `systemctl unmask nftables` BEFORE "
        "`systemctl enable nftables` -- enabling a masked service always "
        "fails.",
    "no_empty_passwords_unix":
        "Do not call bare `pam-auth-update` -- on a system with no "
        "attached terminal it can hang waiting for interactive input. "
        "Editing the /etc/pam.d/* files directly already applies the "
        "change; if you must call pam-auth-update, use "
        "`DEBIAN_FRONTEND=noninteractive pam-auth-update --package`.",
    "grub2_uefi_password":
        "This is a Debian/Ubuntu system, not RHEL/CentOS. The command is "
        "`grub-mkpasswd-pbkdf2` (no '2' after grub) -- `grub2-mkpasswd-pbkdf2` "
        "does not exist here.",
    "accounts_password_pam_unix_authtok":
        "Don't assume /usr/share/pam-configs/cac_unix already has a "
        "'Password:' section formatted like the stock 'unix' profile. `cat` "
        "the file first to see its real current content, and confirm your "
        "sed pattern actually matched something -- sed exits 0 even when it "
        "changes nothing.",
    "accounts_password_pam_unix_no_remember":
        "Don't assume the config file already has a 'Password(-Initial):' "
        "section in the exact form your sed pattern expects. `cat` the file "
        "first and confirm the pattern actually matched -- sed exits 0 even "
        "when it changes nothing.",
    "accounts_password_pam_dcredit":
        "Don't use `sed 's/^dcredit=.../'` assuming that exact no-space "
        "form already exists in /etc/security/pwquality.conf -- if it's "
        "not already there in EXACTLY that form, sed changes nothing and "
        "exits 0 anyway. Check for the key allowing optional whitespace "
        "around the `=` (e.g. `grep -E '^dcredit[[:space:]]*='`) before "
        "deciding whether to edit or append, and use the same tolerant "
        "pattern when editing in place -- don't assume no-space formatting.",
    "accounts_password_pam_ucredit":
        "Check /etc/security/pwquality.conf for an existing `ucredit` line "
        "using a pattern that tolerates whitespace around `=` (e.g. `grep "
        "-E '^ucredit[[:space:]]*='`), not just an exact `^ucredit=` match "
        "-- a line like `ucredit = 0` (spaces around the equals, common in "
        "the file's default commented examples) won't match a no-space "
        "pattern, so you'd wrongly conclude no line exists and append a "
        "second, conflicting one. If an uncommented line already exists in "
        "ANY spacing form, edit it in place -- most tools apply the FIRST "
        "match, so appending a second line at the end has no effect.",
    "accounts_password_pam_minlen":
        "Set this in /etc/security/pwquality.conf (a `minlen=12` line), "
        "the same file the dcredit/ucredit/lcredit/ocredit rules use -- "
        "NOT as an argument on the pam_pwquality.so line in "
        "/etc/pam.d/common-password. Ubuntu's pam_pwquality module reads "
        "its settings from pwquality.conf, and the compliance check reads "
        "that file too; editing common-password's argument list is very "
        "likely checking/changing something the scanner never looks at. "
        "Use the same tolerant `[[:space:]]*=` check-before-edit-or-append "
        "pattern as the other pwquality.conf rules.",
    "accounts_umask_etc_bashrc":
        "Write an actual umask value, e.g. `umask 027` -- not the bare word "
        "`umask` with no argument, which sets nothing.",
    "accounts_tmout":
        "Write ALL THREE required lines directly into the CONTENT of "
        "/etc/profile.d/tmout.sh -- as literal text in the file, not as "
        "commands you merely execute while the remediation script itself "
        "is running (running `readonly TMOUT` / `export TMOUT` in the "
        "remediation script's own shell has no lasting effect -- it only "
        "affects that one throwaway process, not any future login shell). "
        "The file must contain exactly:\n"
        "TMOUT=900\n"
        "readonly TMOUT\n"
        "export TMOUT\n"
        "The compliance check almost certainly greps the FILE for all "
        "three lines, not the live shell state during remediation.",
    "firewall_single_service_active":
        "Don't count active firewall services with a multi-line command "
        "substitution assigned to one variable and compared with `-ne` -- "
        "that breaks bash's numeric test. Check each service individually "
        "with `systemctl is-active <service> --quiet` one at a time, stop "
        "the ones not chosen, then enable exactly one.",
    "file_groupowner_backup_etc_gshadow":
        "After running chgrp, print `ls -l /etc/gshadow-` to confirm the "
        "group change actually persisted and wasn't reset by another "
        "process (e.g. a password-change tool regenerating the backup).",
    "accounts_passwords_pam_faillock_deny":
        "IMPORTANT: do NOT ask the human what number to use here (via "
        "CLARIFY) -- this is a site-tunable policy variable "
        "(var_accounts_passwords_pam_faillock_deny), not a fixed constant, "
        "and the human running this pipeline doesn't know the 'correct' "
        "number any better than you do unless it's written in a tailoring "
        "file you don't have access to. If the reference fix text given to "
        "you doesn't show an already-resolved number, just use `deny = 4` "
        "(Ubuntu's own documented CIS example default) and add a comment "
        "noting this should be confirmed against the real policy value "
        "later. Append to /etc/security/faillock.conf with `>>`, not "
        "overwrite with `>` -- other faillock rules write to this same "
        "file.",
    "accounts_passwords_pam_faillock_unlock_time":
        "IMPORTANT: do NOT ask the human what number to use here (via "
        "CLARIFY) -- this is a site-tunable policy variable, not a fixed "
        "constant, and the human doesn't know the 'correct' number any "
        "better than you do without a tailoring file neither of you has. "
        "If the reference fix text doesn't show an already-resolved "
        "number, just use `unlock_time=900` (15 minutes, a commonly used "
        "default) and add a comment noting this should be confirmed later. "
        "Append to /etc/security/faillock.conf with `>>`, not overwrite "
        "with `>`.",
    "accounts_passwords_pam_faillock_enabled":
        "After creating the new pam-configs profile files, you must "
        "explicitly enable the profile, e.g. `DEBIAN_FRONTEND=noninteractive "
        "pam-auth-update --enable=cac_faillock`. Running bare "
        "`pam-auth-update` does not automatically turn on a newly added "
        "profile.",
    "systemd_journal_upload_server_tls":
        "Don't pass a regex pattern like `[^/]+\\.conf` directly to grep as "
        "if it were a real filename -- use an actual path or a shell glob "
        "loop. Also: commenting out old settings isn't enough on its own -- "
        "you must also write the new required directive's value into the "
        "config file, not just remove old lines.",
    "systemd_journal_upload_url":
        "Don't pass a regex pattern directly to grep as if it were a real "
        "filename -- use an actual path or a shell glob loop. Also: "
        "commenting out old settings isn't enough on its own -- you must "
        "also write the new required URL value into the config file.",
    "sysctl_net_ipv4_tcp_syncookies":
        "Check /etc/sysctl.conf itself for a pre-existing conflicting "
        "`net.ipv4.tcp_syncookies` line before appending a new one to that "
        "same file -- don't only clean up other sysctl.d directories while "
        "leaving a contradictory line in the file you're writing to.",
    "partition_for_tmp":
        "The OpenSCAP check for this rule almost certainly just verifies "
        "/tmp is its own separate mount point (a distinct fstab/mount "
        "entry from /), NOT that it's on a dedicated physical partition. "
        "Do not grab an arbitrary existing block device with lsblk and "
        "mount it onto /tmp -- that can clobber a partition already in use "
        "for something else. Instead, use one of these two standard, safe "
        "approaches: (1) `systemctl unmask tmp.mount && systemctl enable "
        "--now tmp.mount` to turn on Ubuntu's built-in tmpfs unit for /tmp, "
        "or (2) add a tmpfs line to /etc/fstab, e.g. `tmpfs /tmp tmpfs "
        "defaults,rw,nosuid,nodev,noexec,relatime,size=2G 0 0`, then "
        "`mount -o remount /tmp`. Either way this is a real, low-risk fix "
        "-- don't ask to skip it or treat it as impossible.",
}

SYSTEM_INFO = {
    "hostname": "pranjal-garg-VirtualBox",
    "kernel":   "6.17.0-23-generic",
    "os":       "Ubuntu 24.04 LTS (Noble Numbat)",
    "arch":     "x86_64",
}

PREFIX = "xccdf_org.ssgproject.content_rule_"

PROFILES = {
    "1": {
        "key": "MAX",
        "label": "MAX -- System/Cloud Admin, Public Cloud, Production/Critical",
    },
    "2": {
        "key": "MIN",
        "label": "MIN -- Personal Laptop, Just me, Trusted home network only",
    }
}

# ─────────────────────────────────────────────────────────────────────────────
# GROUND TRUTH LOADING (rule-per-row layout, 63 rules, MAX/MIN columns)
# ─────────────────────────────────────────────────────────────────────────────

def load_ground_truth(path):
    """
    Returns an ordered list of dicts, one per rule row:
      {num, category, short, title, max_decision, min_decision,
       conditioned, reason}
    Reads the header row dynamically (looks for '#' in col A) instead of
    hardcoding row numbers, so minor formatting shifts don't break it.
    """
    path = resolve_path(path)
    wb = load_workbook(path, data_only=True)

    ws = None
    for name in GROUND_TRUTH_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
        print(f"  [WARN] No sheet named {GROUND_TRUTH_SHEET_CANDIDATES}; "
              f"using first sheet '{ws.title}' instead.")

    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        if row and row[0] == "#" and (row[1] or "").strip().lower().startswith("categ"):
            header_row_idx = i
            break
    if header_row_idx is None:
        raise RuntimeError(
            f"Could not find header row ('#', 'Category', ...) in sheet "
            f"'{ws.title}'. Check the file structure.")

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or row[0] in (None, "", "Totals") or not isinstance(row[0], (int, float)):
            break  # hit the trailing totals/summary block
        num, category, short, title, max_dec, min_dec = row[0:6]
        conditioned = row[6] if len(row) > 6 else None
        reason      = row[7] if len(row) > 7 else None
        if not short:
            continue
        rows.append({
            "num":         int(num),
            "category":    (category or "").strip(),
            "short":       str(short).strip(),
            "title":       (title or "").strip(),
            "max_decision": (max_dec or "").strip().upper(),
            "min_decision": (min_dec or "").strip().upper(),
            "conditioned": (str(conditioned).strip() if conditioned else ""),
            "reason":      (reason or "").strip(),
        })

    print(f"  Ground truth loaded: {len(rows)} rules from sheet '{ws.title}'")
    return rows


def gt_decision_for_profile(gt_row, profile_key):
    if profile_key == "MAX":
        return gt_row["max_decision"] or "KEEP"
    return gt_row["min_decision"] or "KEEP"


# ─────────────────────────────────────────────────────────────────────────────
# SCAN XML PARSING
# ─────────────────────────────────────────────────────────────────────────────

def parse_scan_xml(path):
    path = resolve_path(path)
    rules = {}
    if not os.path.exists(path):
        print(f"  [WARN] Scan XML '{path}' not found -- rules will run with "
              f"ground-truth titles only (no reference fix / description).")
        return rules
    try:
        tree = ET.parse(path)
    except Exception as e:
        print(f"[ERROR] Cannot parse {path}: {e}")
        return rules

    root = tree.getroot()
    for ns_uri in ["http://checklists.nist.gov/xccdf/1.2",
                   "http://checklists.nist.gov/xccdf/1.1"]:
        for rule_el in root.iter(f"{{{ns_uri}}}Rule"):
            rid = rule_el.get("id", "")
            if not rid.startswith(PREFIX):
                continue

            title_el = rule_el.find(f"{{{ns_uri}}}title")
            title    = (title_el.text or "").strip() if title_el is not None else rid

            desc_el  = rule_el.find(f"{{{ns_uri}}}description")
            desc     = "".join(desc_el.itertext()).strip() if desc_el is not None else ""

            severity = rule_el.get("severity", "unknown")

            fixes = {}
            for fix_el in rule_el.findall(f"{{{ns_uri}}}fix"):
                fixes[fix_el.get("system", "fixtext")] = (fix_el.text or "").strip()
            ft = rule_el.find(f"{{{ns_uri}}}fixtext")
            if ft is not None and ft.text:
                fixes["fixtext"] = ft.text.strip()

            fix_text = (
                fixes.get("urn:xccdf:fix:script:sh")
                or fixes.get("fixtext")
                or next(iter(fixes.values()), "")
            )
            fix_system = (
                "sh"      if "urn:xccdf:fix:script:sh" in fixes else
                "fixtext" if "fixtext" in fixes else
                ("other"  if fixes else "none")
            )

            rules[rid] = {
                "title":       title,
                "description": desc[:1200],
                "fix":         fix_text,
                "fix_system":  fix_system,
                "severity":    severity,
            }

    if not rules:
        print(f"\n  No Rule elements found in {path}.")
        print("  Re-run oscap with --results-arf to embed full benchmark content.")
    return rules


# ─────────────────────────────────────────────────────────────────────────────
# GROUND-TRUTH SHORT NAME  <->  OFFICIAL XCCDF RULE ID FUZZY MATCHING
# ─────────────────────────────────────────────────────────────────────────────

def _norm(s):
    return s.replace("-", "_").lower().strip("_")


def resolve_rule_id(gt_short, scan_rules):
    """
    Returns (rule_id, matched_bool).
    matched_bool True  -> rule_id is a real key in scan_rules (full metadata available)
    matched_bool False -> best-guess id (PREFIX + gt_short); rule still runs,
                          just without scan-XML title/description/fix/severity.
    """
    guess = PREFIX + gt_short
    if guess in scan_rules:
        return guess, True

    gt_n = _norm(gt_short)
    for full_id in scan_rules:
        short = full_id[len(PREFIX):] if full_id.startswith(PREFIX) else full_id
        short_n = _norm(short)
        if short_n == gt_n or short_n.endswith("_" + gt_n) or short_n.endswith(gt_n):
            return full_id, True

    # looser fallback: gt short appears anywhere in an official short name
    for full_id in scan_rules:
        short = full_id[len(PREFIX):] if full_id.startswith(PREFIX) else full_id
        if gt_n in _norm(short):
            return full_id, True

    return guess, False


def build_rule_set(gt_rows, scan_rules):
    """
    Produces the ordered list of ALL 63 rules, each entry:
      {gt, rule_id, matched, info}
    'info' has title/description/fix/severity -- from scan XML if matched,
    else falls back to the ground truth title only.

    NOTE: this still returns ALL rules (KEEP and SKIP). Filtering by
    ground-truth decision for the chosen profile happens later, per-profile,
    in run_one_model() -- because MAX and MIN can have different decisions
    for the same rule, so filtering here (before profile is even known in
    some code paths) would be wrong.
    """
    resolved = []
    unmatched = []
    for gt in gt_rows:
        rule_id, matched = resolve_rule_id(gt["short"], scan_rules)
        if matched:
            info = scan_rules[rule_id]
        else:
            unmatched.append(gt["short"])
            info = {
                "title": gt["title"] or gt["short"],
                "description": f"(Not found in scan XML -- CIS category: {gt['category']}. "
                                f"Apply the standard Ubuntu 24.04 remediation for this control.)",
                "fix": "",
                "fix_system": "none",
                "severity": "unknown",
            }
        resolved.append({"gt": gt, "rule_id": rule_id, "matched": matched, "info": info})

    print(f"  Rule set built: {len(resolved)} total "
          f"({len(resolved) - len(unmatched)} matched to scan XML, "
          f"{len(unmatched)} using ground-truth title only)")
    if unmatched:
        print(f"    Unmatched (still WILL run if KEEP): {unmatched}")
    return resolved


def filter_to_min_skipped(rule_set):
    """
    Restricts rule_set to only the rules whose MIN-profile ground-truth
    decision was SKIP. Used for --only-min-delta: when you've already fully
    tested all KEEP-under-MIN rules for a profile, and MAX's KEEP set is a
    superset of MIN's KEEP set, the only new ground to cover for MAX is the
    rules MIN skipped. Everything else (KEEP-under-MIN rules) is assumed
    already covered by a prior MIN run and is excluded here entirely --
    not just marked skipped_by_ground_truth, but removed from the rule_set
    before run_one_model ever sees it, so it doesn't even print/log a line
    for those 51.
    """
    filtered = [item for item in rule_set if item["gt"]["min_decision"] != "KEEP"]
    excluded_shorts = [item["gt"]["short"] for item in rule_set
                        if item["gt"]["min_decision"] == "KEEP"]
    print(f"\n  [--only-min-delta] Restricting to rules where MIN decision "
          f"was SKIP: {len(filtered)} of {len(rule_set)} rules kept.")
    print(f"    Excluded (already covered by prior MIN run): {len(excluded_shorts)} rules")
    for item in filtered:
        gt = item["gt"]
        print(f"    - {gt['short']:45s} MIN={gt['min_decision']:5s} MAX={gt['max_decision']}")
    return filtered


def filter_to_only_rules(rule_set, wanted):
    """Restricts rule_set to items whose rule_id or short name is in
    `wanted` (a set of strings, already normalized to short-name form)."""
    def short_of(rule_id):
        return rule_id[len(PREFIX):] if rule_id.startswith(PREFIX) else rule_id

    filtered = [item for item in rule_set if short_of(item["rule_id"]) in wanted]
    found = {short_of(item["rule_id"]) for item in filtered}
    missing = wanted - found
    if missing:
        print(f"  [WARN] --only-rules: not found in rule set, ignored: {sorted(missing)}")
    print(f"  [--only-rules] Restricting to {len(filtered)} of {len(rule_set)} rules: "
          f"{sorted(found)}")
    return filtered


def filter_to_retry_failed(rule_set, json_path, profile_key):
    """Restricts rule_set to rules that were NOT oscap_pass (and not
    skipped_by_ground_truth) in a prior result JSON for this profile --
    i.e. only what still needs fixing. Anything the prior run never touched
    (e.g. it used --only-rules itself) is left OUT, since there's nothing to
    say it needs a retry."""
    with open(json_path) as f:
        prior = json.load(f)
    if prior.get("profile") and prior["profile"] != profile_key:
        print(f"  [WARN] --retry-failed file was run against profile "
              f"'{prior.get('profile')}', not '{profile_key}' -- statuses may "
              f"not line up as expected.")

    prior_rules = prior.get("rules", {})
    needs_retry = {
        rid for rid, rec in prior_rules.items()
        if rec.get("status") not in ("oscap_pass", "skipped_by_ground_truth")
    }

    def short_of(rule_id):
        return rule_id[len(PREFIX):] if rule_id.startswith(PREFIX) else rule_id

    filtered = [item for item in rule_set if item["rule_id"] in needs_retry]
    print(f"  [--retry-failed] {json_path}: {len(needs_retry)} rules were not "
          f"oscap_pass -> restricting to {len(filtered)} of {len(rule_set)} rules: "
          f"{sorted(short_of(item['rule_id']) for item in filtered)}")
    return filtered


# ─────────────────────────────────────────────────────────────────────────────
# PROFILE SELECTION
# ─────────────────────────────────────────────────────────────────────────────

def ask_profile():
    print("\n" + "=" * 60)
    print("  Select ground-truth profile to test against")
    print("=" * 60)
    for k, v in PROFILES.items():
        print(f"  {k}. {v['label']}")
    while True:
        c = input("\nProfile number: ").strip()
        if c in PROFILES:
            return PROFILES[c]["key"]
        print("Invalid, try again.")


# ─────────────────────────────────────────────────────────────────────────────
# LAB SERVER QUERY  (with multi-endpoint fallback + diagnostics)
# ─────────────────────────────────────────────────────────────────────────────

def _try_openai_chat(model, prompt, max_tokens, timeout):
    url = f"{LAB_URL}/v1/chat/completions"
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
        "temperature": 0.2,
        "stream": False,
    }
    resp = requests.post(url, json=payload, auth=(LAB_USER, LAB_PASS),
                          verify=False, timeout=timeout)
    return resp, url


def _try_ollama_chat(model, prompt, max_tokens, timeout):
    url = f"{LAB_URL}/api/chat"
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "stream": False,
        "options": {"temperature": 0.2, "num_predict": max_tokens},
    }
    resp = requests.post(url, json=payload, auth=(LAB_USER, LAB_PASS),
                          verify=False, timeout=timeout)
    return resp, url


def _try_ollama_generate(model, prompt, max_tokens, timeout):
    url = f"{LAB_URL}/api/generate"
    payload = {
        "model": model,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": 0.2, "num_predict": max_tokens},
    }
    resp = requests.post(url, json=payload, auth=(LAB_USER, LAB_PASS),
                          verify=False, timeout=timeout)
    return resp, url


_ENDPOINT_TRIED_ORDER = [_try_openai_chat, _try_ollama_chat, _try_ollama_generate]
_WORKING_ENDPOINT = [None]  # cache which shape worked, so we don't re-probe every call


def query_lab_model(model_name, prompt, batch_len=20, timeout=900):
    """
    Queries LAB_URL/chat and extracts the reply text regardless of which
    JSON shape the proxy returns (OpenAI 'choices', custom 'message', or
    plain 'response'). The shape-detection happens BEFORE any dict access
    that could raise KeyError, so an unexpected shape degrades gracefully
    into a debug printout instead of crashing the whole rule.
    """
    url = f"{LAB_URL}/chat"
    payload = {
        "model": model_name,
        "messages": [{"role": "user", "content": prompt}],
    }

    try:
        response = requests.post(
            url,
            auth=(LAB_USER, LAB_PASS),
            json=payload,
            verify=False,
            timeout=timeout,
        )
        response.raise_for_status()
        data = response.json()
    except Exception as e:
        return {"model": model_name, "response": "", "error": str(e), "fatal": True}

    reply_text = None
    try:
        if "choices" in data:
            reply_text = data["choices"][0]["message"]["content"].strip()
        elif "message" in data:
            # Some custom proxies return {"message": {"role": "...", "content": "..."}}
            msg = data["message"]
            reply_text = (msg.get("content", "") if isinstance(msg, dict) else str(msg)).strip()
        elif "response" in data:
            # Some return {"response": "..."}
            reply_text = str(data["response"]).strip()
        else:
            print(f"    [DEBUG] Unexpected JSON structure. Keys: {list(data.keys())}")
            reply_text = str(data)
    except Exception as e:
        print(f"    [DEBUG] Error extracting content: {e}. Data: {data}")
        return {
            "model": model_name, "response": "",
            "error": f"content extraction failed: {e}", "fatal": True,
        }

    if not reply_text:
        return {
            "model": model_name, "response": "",
            "error": "empty response body after extraction", "fatal": True,
        }

    return {
        "model": model_name,
        "response": reply_text,
        "elapsed_seconds": 0,
        "error": None,
        "fatal": False,
    }


def probe_lab_server():
    """--probe mode: hit several endpoints/paths and print raw results so you
    can see exactly what your lab server exposes."""
    print("\n" + "=" * 60)
    print(f"  PROBING LAB SERVER: {LAB_URL}")
    print("=" * 60)

    checks = [
        ("GET",  "/v1/models"),
        ("GET",  "/api/tags"),
        ("GET",  "/"),
        ("GET",  "/health"),
    ]
    for method, path in checks:
        url = LAB_URL + path
        try:
            resp = requests.request(method, url, auth=(LAB_USER, LAB_PASS),
                                     verify=False, timeout=15)
            print(f"  {method} {path:20s} -> {resp.status_code}  {resp.text[:150]!r}")
        except requests.exceptions.RequestException as e:
            print(f"  {method} {path:20s} -> ERROR: {e}")

    test_model = MODELS[0]
    print(f"\n  Now trying actual chat-completion shapes with model '{test_model}':")
    for fn in _ENDPOINT_TRIED_ORDER:
        try:
            resp, url = fn(test_model, "Say OK.", 20, 30)
            print(f"  {fn.__name__:20s} [{url}] -> {resp.status_code}  {resp.text[:200]!r}")
        except requests.exceptions.RequestException as e:
            print(f"  {fn.__name__:20s} -> ERROR: {e}")

    print("\n  Whichever shape/path returned 200 above with a real completion is")
    print("  the one to keep; tell me the working path and I'll hardcode it so")
    print("  every call skips straight to it instead of probing.")


def strip_code_fences(text):
    text = re.sub(r"^```[a-zA-Z]*\n?", "", text.strip())
    text = re.sub(r"\n?```$", "", text)
    return text.strip()


# ─────────────────────────────────────────────────────────────────────────────
# PROMPT BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_prompt(rule_id, rule_info, profile_key, gt_row):
    style_note = "Give the fix as a runnable bash script (it will be run directly with `sudo bash -c`)."

    gt_decision = gt_decision_for_profile(gt_row, profile_key)
    gt_reason   = gt_row.get("reason") or "(no reason recorded)"

    profile_desc = (
        "System/Cloud Administrator running a production, internet-facing "
        "workload in a public cloud"
        if profile_key == "MAX" else
        "individual on a personal laptop, used only by them, on a trusted "
        "home network"
    )

    short_id = rule_id[len(PREFIX):] if rule_id.startswith(PREFIX) else rule_id
    hint = RULE_HINTS.get(short_id)
    hint_block = (
        f"\nKNOWN ISSUE FOR THIS SPECIFIC RULE (from a prior run on this "
        f"exact host) -- read carefully and account for it:\n{hint}\n"
        if hint else ""
    )

    known_fix = KNOWN_FIXES.get(short_id)
    known_fix_block = (
        f"\nVERIFIED WORKING SCRIPT FOR THIS EXACT RULE ON THIS EXACT HOST "
        f"-- this has already been tested and confirmed correct on this "
        f"system. Use it AS-IS. Only deviate if you can see a concrete, "
        f"specific reason it won't work (e.g. it references a package that "
        f"genuinely isn't installed here) -- do not rewrite it from scratch "
        f"or 'improve' it without a real reason:\n```bash\n{known_fix}\n```\n"
        if known_fix else ""
    )

    return f"""You are a Linux system hardening expert. Write a remediation script for
the following CIS/OpenSCAP rule on this specific host.

Host: {SYSTEM_INFO['hostname']} | {SYSTEM_INFO['os']} | kernel {SYSTEM_INFO['kernel']} | {SYSTEM_INFO['arch']}
Profile: {profile_key} -- {profile_desc}

Rule ID:     {rule_id}
Title:       {rule_info['title']}
Severity:    {rule_info['severity']}
Description: {rule_info['description']}

Ground-truth expert assessment for this profile:
  Decision: {gt_decision}
  Reason:   {gt_reason}
(This rule is only sent to you because the ground-truth decision for this
profile is KEEP -- write the correct technical remediation script for it.)

Reference fix from benchmark (format: {rule_info.get('fix_system', 'none')}):
{rule_info['fix'] or '(none provided -- write the standard Ubuntu 24.04 remediation yourself)'}

If the reference fix is bash (sh), adapt it directly.
If it is Ansible/Puppet/blueprint or missing, translate the intent into plain bash.
{hint_block}
{known_fix_block}
The script will be executed with `sudo bash -c "<script>"`, so:
  - Do NOT prefix every individual line with `sudo` (it's redundant and can
    break heredocs/pipes) -- the whole script already runs as root.
  - Just write plain root-level commands.

{style_note}
Output ONLY the script -- no prose, no markdown fences, no explanation.

If -- and only if -- you have a genuine, non-technical doubt about what's
actually wanted here (e.g. the rule and the stated profile intent seem to
point in different directions, or the reference fix is ambiguous about
scope), you may instead ask a human ONE clarifying question, in exactly
this format and nothing else:
CLARIFY
QUESTION: <your question, plain language, no jargon>
OPTIONS:
1) <option>
2) <option>
3) <option>
Do not use this for ordinary technical uncertainty (missing packages,
unfamiliar paths, etc) -- just write your best script for those and let the
error tell you if you're wrong.
"""


# ─────────────────────────────────────────────────────────────────────────────
# OSCAP SINGLE-RULE VERIFICATION
# ─────────────────────────────────────────────────────────────────────────────

def verify_rule(rule_id):
    cmd = (
        f"sudo oscap xccdf eval --rule {rule_id} "
        f"{BENCHMARK_XML} 2>/dev/null | grep 'Result'"
    )
    try:
        out = subprocess.check_output(
            cmd, shell=True, text=True, timeout=90
        ).strip().lower()
        if "pass" in out:            return "pass"
        elif "fail" in out:          return "fail"
        elif "notapplicable" in out: return "notapplicable"
        else:                        return "notchecked"
    except subprocess.TimeoutExpired:
        return "timeout"
    except Exception:
        return "error"


# ─────────────────────────────────────────────────────────────────────────────
# BREAK-RULES SCRIPT GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

BREAK_COMMANDS = {
    "accounts_passwords_pam_faillock_deny":
        "sudo sed -i '/faillock/d' /etc/pam.d/common-auth /etc/pam.d/common-account 2>/dev/null || true",
    "accounts_passwords_pam_faillock_enabled":
        "sudo sed -i '/pam_faillock/d' /etc/pam.d/common-auth /etc/pam.d/common-account 2>/dev/null || true",
    "accounts_passwords_pam_faillock_unlock_time":
        "sudo sed -i '/faillock/d' /etc/pam.d/common-auth 2>/dev/null || true",
    "accounts_password_pam_dcredit":
        "sudo sed -i '/dcredit/d' /etc/security/pwquality.conf 2>/dev/null || true",
    "accounts_password_pam_minlen":
        "sudo sed -i '/minlen/d' /etc/security/pwquality.conf 2>/dev/null || true",
    "accounts_password_pam_ucredit":
        "sudo sed -i '/ucredit/d' /etc/security/pwquality.conf 2>/dev/null || true",
    "accounts_password_pam_unix_no_remember":
        "sudo sed -i '/remember/d' /etc/pam.d/common-password 2>/dev/null || true",
    "accounts_password_pam_unix_authtok":
        "sudo sed -i 's/ *sha512//' /etc/pam.d/common-password 2>/dev/null || true",
    "set_password_hashing_algorithm_systemauth":
        "sudo sed -i 's/ *sha512//' /etc/pam.d/common-password 2>/dev/null || true",
    "no_empty_passwords_unix":
        "sudo sed -i 's/ nullok_secure//g' /etc/pam.d/common-auth 2>/dev/null || true",
    "accounts_tmout":
        "sudo sed -i '/TMOUT/d' /etc/bash.bashrc /etc/profile /etc/profile.d/*.sh 2>/dev/null || true",
    "accounts_umask_etc_bashrc":
        "sudo sed -i '/umask 027/d; /umask 077/d' /etc/bash.bashrc 2>/dev/null || true",
    "umask_etc_login_defs":
        "sudo sed -i 's/^UMASK.*/UMASK 022/' /etc/login.defs 2>/dev/null || true",
    "sudo_custom_logfile":
        "sudo sed -i '/logfile/d' /etc/sudoers 2>/dev/null; sudo rm -f /etc/sudoers.d/logfile 2>/dev/null || true",
    "sudo_require_reauthentication":
        "sudo rm -f /etc/sudoers.d/timeout 2>/dev/null; echo 'Defaults timestamp_timeout=15' | sudo tee /etc/sudoers.d/noauth > /dev/null",
    "sudo_remove_no_authenticate":
        "echo 'Defaults !authenticate' | sudo tee /etc/sudoers.d/zz_break_authenticate > /dev/null",
    "package_apparmor-utils_installed":
        "sudo apt-get remove -y apparmor-utils 2>/dev/null || true",
    "grub2_enable_apparmor":
        "sudo sed -i 's/ apparmor=1 security=apparmor//' /etc/default/grub 2>/dev/null; sudo update-grub 2>/dev/null || true",
    "all_apparmor_profiles_in_enforce_complain_mode":
        "sudo aa-complain /etc/apparmor.d/* 2>/dev/null || true",
    "sysctl_net_ipv6_conf_all_forwarding":
        "sudo sed -i '/net.ipv6.conf.all.forwarding/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv6.conf.all.forwarding=1 2>/dev/null || true",
    "sysctl_net_ipv4_conf_all_accept_redirects":
        "sudo sed -i '/accept_redirects/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.conf.all.accept_redirects=1 2>/dev/null || true",
    "sysctl_net_ipv4_conf_all_log_martians":
        "sudo sed -i '/log_martians/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.conf.all.log_martians=0 2>/dev/null || true",
    "sysctl_net_ipv4_conf_all_rp_filter":
        "sudo sed -i '/rp_filter/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.conf.all.rp_filter=0 2>/dev/null || true",
    "sysctl_net_ipv4_tcp_syncookies":
        "sudo sed -i '/syncookies/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.tcp_syncookies=0 2>/dev/null || true",
    "sysctl_net_ipv4_conf_all_send_redirects":
        "sudo sed -i '/send_redirects/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.conf.all.send_redirects=1 2>/dev/null || true",
    "sysctl_net_ipv4_ip_forward":
        "sudo sed -i '/ip_forward/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.ip_forward=1 2>/dev/null || true",
    "sysctl_kernel_randomize_va_space":
        "sudo sed -i '/randomize_va_space/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w kernel.randomize_va_space=0 2>/dev/null || true",
    "sysctl_fs_suid_dumpable":
        "sudo sed -i '/suid_dumpable/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w fs.suid_dumpable=1 2>/dev/null || true",
    "disable_users_coredumps":
        "sudo sed -i '/hard core/d; /soft core/d' /etc/security/limits.conf 2>/dev/null; echo '* soft core unlimited' | sudo tee -a /etc/security/limits.conf",
    "kernel_module_cramfs_disabled":
        "sudo rm -f /etc/modprobe.d/cramfs.conf 2>/dev/null || true",
    "kernel_module_hfs_disabled":
        "sudo rm -f /etc/modprobe.d/hfs.conf 2>/dev/null || true",
    "kernel_module_hfsplus_disabled":
        "sudo rm -f /etc/modprobe.d/hfsplus.conf 2>/dev/null || true",
    "kernel_module_jffs2_disabled":
        "sudo rm -f /etc/modprobe.d/jffs2.conf 2>/dev/null || true",
    "mount_option_dev_shm_nodev":
        "sudo sed -i '/\\/dev\\/shm/d' /etc/fstab; sudo mount -o remount /dev/shm 2>/dev/null || true",
    "mount_option_dev_shm_noexec":
        "sudo sed -i '/\\/dev\\/shm/d' /etc/fstab; sudo mount -o remount /dev/shm 2>/dev/null || true",
    "mount_option_dev_shm_nosuid":
        "sudo sed -i '/\\/dev\\/shm/d' /etc/fstab; sudo mount -o remount /dev/shm 2>/dev/null || true",
    "file_permissions_crontab":
        "sudo chmod 0644 /etc/crontab 2>/dev/null || true",
    "file_permissions_cron_allow":
        "sudo chmod 0644 /etc/cron.allow 2>/dev/null || true",
    "file_permissions_cron_d":
        "sudo chmod 0755 /etc/cron.d 2>/dev/null || true",
    "file_permissions_cron_daily":
        "sudo chmod 0755 /etc/cron.daily 2>/dev/null || true",
    "file_owner_cron_allow":
        "sudo chown nobody:nogroup /etc/cron.allow 2>/dev/null || true",
    "file_groupowner_cron_allow":
        "sudo chgrp root /etc/cron.allow 2>/dev/null || true",
    "file_groupowner_backup_etc_gshadow":
        "sudo chgrp root /etc/gshadow- 2>/dev/null || true",
    "package_ftp_removed":
        "sudo apt-get install -y ftp 2>/dev/null || true",
    "package_tnftp_removed":
        "sudo apt-get install -y tnftp 2>/dev/null || true",
    "package_openldap-clients_removed":
        "sudo apt-get install -y ldap-utils 2>/dev/null || true",
    "package_rsync_removed":
        "sudo apt-get install -y rsync 2>/dev/null || true",
    "service_rsyncd_disabled":
        "sudo systemctl enable rsync 2>/dev/null || sudo systemctl enable rsyncd 2>/dev/null || true",
    "package_telnet_removed":
        "sudo apt-get install -y telnet 2>/dev/null || true",
    "package_vsftpd_removed":
        "sudo apt-get install -y vsftpd 2>/dev/null || true",
    "service_vsftpd_disabled":
        "sudo systemctl enable vsftpd 2>/dev/null || true",
    "package_nis_removed":
        "sudo apt-get install -y nis 2>/dev/null || true",
    "package_rpcbind_removed":
        "sudo apt-get install -y rpcbind 2>/dev/null || true",
    "package_ypserv_removed":
        "sudo apt-get install -y nis 2>/dev/null || true",
}


def generate_break_script(rule_set, profile_key, output_path="break_rules.sh"):
    """
    Only generates break commands for rules whose ground-truth decision for
    this profile is KEEP -- SKIP rules are never remediated, so there's
    nothing to break/reset for them.
    """
    not_found = []
    lines = [
        "#!/bin/bash",
        "# Auto-generated by remediationv3.py",
        "# Resets KEEP-category rules back to FAILING state between models",
        "# Run between models: bash break_rules.sh",
        "echo '=== Resetting rules to failing state ==='",
        "",
    ]
    for item in rule_set:
        if gt_decision_for_profile(item["gt"], profile_key) != "KEEP":
            continue
        short = item["rule_id"].replace(PREFIX, "")
        gt_short = item["gt"]["short"]
        cmd = BREAK_COMMANDS.get(short) or BREAK_COMMANDS.get(gt_short)
        if cmd:
            lines.append(f"echo '  Breaking: {short}'")
            lines.append(cmd)
        else:
            not_found.append(short)

    lines += ["", "echo '=== Reset complete ==='"]
    if not_found:
        lines.append(f"echo 'No break command for: {not_found}'")

    with open(output_path, "w") as f:
        f.write("\n".join(lines) + "\n")
    os.chmod(output_path, 0o755)

    print(f"\n  Break script written to: {output_path}")
    if not_found:
        print(f"  KEEP rules without break command (add manually): {not_found}")
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# SCRIPT EXECUTION (elevated)
# ─────────────────────────────────────────────────────────────────────────────

def run_remediation_script(script, timeout=120):
    """
    Executes the LLM-generated remediation script with root privileges.
    Fixes the earlier 'Permission denied' failures (apt-get, aideinit,
    /etc/sudoers.d writes, etc.) which happened because scripts were run as
    an unprivileged user via bare `bash -c`.
    """
    if ELEVATE_WITH_SUDO:
        # -n = non-interactive: if a password would be required, sudo fails
        # immediately with a clear error instead of hanging on stdin (which,
        # with no TTY attached under subprocess, previously caused every
        # script to silently block until the 120s timeout killed it).
        cmd = ["sudo", "-n", "bash", "-c", script]
    else:
        cmd = ["bash", "-c", script]
    return subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)


def run_remediation_script_interactive(script, timeout=180):
    """Like run_remediation_script, but connects to the real terminal instead
    of capturing output, so `sudo` can prompt for (and receive) a password
    directly. Only called after the human explicitly grants permission via
    ask_permission() -- never on its own."""
    cmd = ["sudo", "bash", "-c", script]
    return subprocess.run(cmd, timeout=timeout)


# ─────────────────────────────────────────────────────────────────────────────
# MULTI-PASS AGENT LOOP (v3.3)
# ─────────────────────────────────────────────────────────────────────────────
# Each KEEP rule gets up to MAX_PASSES attempts instead of one shot. Between
# attempts: script errors and oscap failures are fed back to the model so it
# can revise its own script; a sudoers write is validated and rolled back if
# it breaks `sudo` for later rules; a "needs a password" failure pauses for
# explicit human permission instead of quietly failing.

SUDOERS_TOUCHING_HINTS = ("sudoers", "visudo")

_PASSWORD_REQUIRED_PATTERNS = (
    "a password is required",
    "sudo: a terminal is required",
    "sudo: no tty present",
)


def script_touches_sudoers(script):
    low = script.lower()
    return any(h in low for h in SUDOERS_TOUCHING_HINTS)


def snapshot_sudoers():
    """Backs up /etc/sudoers and /etc/sudoers.d before a risky script runs.
    Returns the backup dir, or None if the snapshot itself failed (rollback
    is then skipped for this attempt rather than trusting a bad backup)."""
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    backup_dir = f"/tmp/sudoers_backup_{ts}"
    try:
        subprocess.run(["sudo", "mkdir", "-p", backup_dir],
                        check=True, capture_output=True, timeout=15)
        subprocess.run(["sudo", "cp", "-a", "/etc/sudoers", f"{backup_dir}/sudoers"],
                        check=True, capture_output=True, timeout=15)
        subprocess.run(["sudo", "cp", "-a", "/etc/sudoers.d", f"{backup_dir}/sudoers.d"],
                        check=True, capture_output=True, timeout=15)
        return backup_dir
    except Exception as e:
        print(f"    [WARN] Could not snapshot sudoers, rollback disabled this pass: {e}")
        return None


def sudoers_is_valid():
    try:
        proc = subprocess.run(["sudo", "-n", "visudo", "-c"],
                               capture_output=True, text=True, timeout=15)
        return proc.returncode == 0, (proc.stdout + proc.stderr).strip()
    except Exception as e:
        return False, str(e)


def restore_sudoers(backup_dir):
    if not backup_dir:
        return False
    try:
        subprocess.run(["sudo", "cp", "-a", f"{backup_dir}/sudoers", "/etc/sudoers"],
                        check=True, capture_output=True, timeout=15)
        subprocess.run(["sudo", "rm", "-rf", "/etc/sudoers.d"],
                        check=True, capture_output=True, timeout=15)
        subprocess.run(["sudo", "cp", "-a", f"{backup_dir}/sudoers.d", "/etc/sudoers.d"],
                        check=True, capture_output=True, timeout=15)
        return True
    except Exception as e:
        print(f"    [ERROR] sudoers rollback failed: {e} -- fix /etc/sudoers "
              f"manually with `visudo` before continuing!")
        return False


def detect_password_gate(stderr):
    low = (stderr or "").lower()
    return any(p in low for p in _PASSWORD_REQUIRED_PATTERNS)


def ask_permission(rule_short, reason):
    """Human-in-the-loop gate for anything the agent can't push through on
    its own. Returns True/False."""
    print(f"\n    [PERMISSION NEEDED] {rule_short}")
    print(f"    Reason: {reason}")
    ans = input("    Allow the agent to proceed with this? [y/n]: ").strip().lower()
    return ans == "y"


def parse_model_output(raw):
    """Splits a model reply into a CLARIFY request or a script. Anything
    that doesn't start with a bare 'CLARIFY' line is treated as a script."""
    stripped = raw.strip()
    first_line = stripped.split("\n", 1)[0].strip().upper()
    if first_line != "CLARIFY":
        return {"type": "script", "script": strip_code_fences(raw)}

    question = ""
    options = []
    q_match = re.search(r"QUESTION:\s*(.+)", stripped)
    if q_match:
        question = q_match.group(1).split("\n")[0].strip()
    for line in stripped.split("\n"):
        m = re.match(r"\s*\d+\)\s*(.+)", line)
        if m:
            options.append(m.group(1).strip())

    if not question:
        # Said CLARIFY but gave nothing usable -- fall back to treating the
        # whole reply as a (probably broken) script rather than blocking.
        return {"type": "script", "script": strip_code_fences(raw)}
    return {"type": "clarify", "question": question, "options": options}


def ask_user_clarification(rule_short, question, options):
    print(f"\n    [AGENT HAS A QUESTION] {rule_short}")
    print(f"    {question}")
    for i, opt in enumerate(options, 1):
        print(f"      {i}) {opt}")
    print(f"      0) Something else (type your own answer)")
    choice = input("    Your answer [number, or 0 for free text]: ").strip()
    valid_choices = [str(i) for i in range(1, len(options) + 1)]
    if choice == "0" or choice not in valid_choices:
        return input("    Type your answer: ").strip()
    return options[int(choice) - 1]


def build_retry_prompt(base_prompt, attempt_history, clarification=None):
    """Appends failure context from prior passes (and any human
    clarification) onto the original prompt so the model sees exactly what
    it tried and why it failed, instead of guessing blind again."""
    parts = [base_prompt]
    if clarification:
        parts.append(f"\nThe human clarified: {clarification}\n"
                      f"Use this to inform your script. Do not ask the same "
                      f"question again.")
    for i, att in enumerate(attempt_history, 1):
        parts.append(f"\n--- Your attempt #{i} failed ---")
        parts.append(f"Script you wrote:\n{att['script']}")
        parts.append(f"Failure reason: {att['failure_reason']}")
        if att.get("detail"):
            parts.append(f"Details:\n{str(att['detail'])[:1000]}")
    parts.append(
        "\nWrite a corrected script that avoids the above problem(s). If "
        "you still have a genuine doubt, use the CLARIFY format instead. "
        "Output ONLY the script or ONLY the CLARIFY block -- no prose, no "
        "markdown fences."
    )
    return "\n".join(parts)


def remediate_rule(model, rule_id, rule_info, gt_row, decision, profile_key,
                    auto_approve, matched):
    """
    Multi-pass remediation for one KEEP rule: query -> apply -> verify,
    retrying with error feedback (and at most one human clarification) up to
    MAX_PASSES times. Returns (rule_record, counters) where counters is a
    dict of summary-key -> increment for the caller to merge in.
    """
    short = rule_id.replace(PREFIX, "")
    base_prompt = build_prompt(rule_id, rule_info, profile_key, gt_row)

    counters = {}
    attempt_history = []
    clarification = None
    clarifications_used = 0
    passes_log = []

    rule_record = {
        "matched_in_scan": matched,
        "gt_decision": decision,
        "gt_reason": gt_row.get("reason", ""),
    }

    pass_num = 0
    while pass_num < MAX_PASSES:
        pass_num += 1
        prompt = base_prompt if (pass_num == 1 and not clarification) \
            else build_retry_prompt(base_prompt, attempt_history, clarification)

        print(f"    Querying {model} (pass {pass_num}/{MAX_PASSES})...")
        try:
            result_dict = query_lab_model(model, prompt)
            if result_dict.get("error"):
                raise Exception(result_dict["error"])
            raw = result_dict["response"]
        except Exception as e:
            print(f"    [QUERY ERROR] {e}")
            rule_record["status"] = "query_error"
            rule_record["error"]  = str(e)
            rule_record["passes"] = passes_log
            counters["query_error"] = 1
            time.sleep(2)
            return rule_record, counters

        parsed = parse_model_output(raw)

        if parsed["type"] == "clarify":
            if clarifications_used >= 1:
                # Already used our one clarification -- treat a repeat
                # question as a failed attempt so the loop still terminates.
                print(f"    [IGNORING REPEAT QUESTION] Model asked again instead "
                      f"of writing a script: \"{parsed['question']}\" -- "
                      f"already clarified once for this rule, treating as a "
                      f"failed attempt.")
                attempt_history.append({
                    "script": "(model asked another question instead of a script)",
                    "failure_reason": "repeated_clarify_ignored",
                    "detail": parsed["question"],
                })
                passes_log.append({"pass": pass_num, "type": "clarify_ignored",
                                    "question": parsed["question"]})
                continue
            clarification = ask_user_clarification(short, parsed["question"], parsed["options"])
            clarifications_used += 1
            counters["clarifications_asked"] = counters.get("clarifications_asked", 0) + 1
            passes_log.append({"pass": pass_num, "type": "clarify",
                                "question": parsed["question"], "answer": clarification})
            pass_num -= 1  # doesn't consume a script-attempt slot
            continue

        script = parsed["script"]
        if "attempted" not in counters:
            counters["attempted"] = 1

        print("\n    Proposed fix:")
        print("    " + "-" * 52)
        for line in script.split("\n")[:20]:
            print(f"    {line}")
        if script.count("\n") > 20:
            print(f"    ... ({script.count(chr(10)) - 20} more lines)")
        print("    " + "-" * 52)

        if pass_num == 1:
            if auto_approve:
                approved = True
                print("    [AUTO] Applying.")
            else:
                ans = input("    Apply? [y/n/s=show full]: ").strip().lower()
                if ans == "s":
                    print(f"\n{script}\n")
                    ans = input("    Apply? [y/n]: ").strip().lower()
                approved = (ans == "y")
            rule_record["approved"] = approved
            if not approved:
                rule_record["script"] = script
                rule_record["status"] = "rejected"
                rule_record["passes"] = passes_log
                counters["rejected"] = 1
                return rule_record, counters
            counters["approved"] = 1
        else:
            print("    [AUTO-RETRY] Applying revised script (rule already approved).")

        rule_record["script"] = script  # most recent attempt

        backup_dir = snapshot_sudoers() if script_touches_sudoers(script) else None
        timeout_s = RULE_TIMEOUTS.get(short, DEFAULT_SCRIPT_TIMEOUT)

        try:
            proc = run_remediation_script(script, timeout=timeout_s)
            exit_code, out, err = proc.returncode, proc.stdout, proc.stderr
        except subprocess.TimeoutExpired:
            print(f"    [TIMEOUT] Script did not finish within {timeout_s}s.")
            passes_log.append({"pass": pass_num, "type": "script", "status": "timeout"})
            attempt_history.append({"script": script, "failure_reason": "timeout",
                                     "detail": f"Script did not finish within {timeout_s}s."})
            continue
        except Exception as e:
            print(f"    [RUNTIME ERROR] {e}")
            passes_log.append({"pass": pass_num, "type": "script", "status": "error",
                                "error": str(e)})
            attempt_history.append({"script": script, "failure_reason": "runtime_error",
                                     "detail": str(e)})
            continue

        rule_record["exit_code"] = exit_code
        rule_record["stdout"]    = (out or "")[-2000:]
        rule_record["stderr"]    = (err or "")[-2000:]

        # -- permission gate: interactive sudo password needed --
        if exit_code != 0 and detect_password_gate(err):
            allowed = ask_permission(
                short,
                "The script needs an interactive sudo password (no NOPASSWD "
                "rule covers this command).")
            if allowed:
                try:
                    proc2 = run_remediation_script_interactive(script, timeout=timeout_s)
                    exit_code = proc2.returncode
                    # proc2's output wasn't captured (it went straight to the
                    # real terminal so the password prompt could work) --
                    # overwrite err/out so nothing below still shows the
                    # stale "password is required" text from the earlier,
                    # non-interactive attempt.
                    out = "(ran interactively, output not captured)"
                    err = "" if exit_code == 0 else \
                        "(ran interactively -- see terminal output above " \
                        "for the actual error, not the password message)"
                    rule_record["exit_code"] = exit_code
                    rule_record["stdout"] = out
                    rule_record["stderr"] = err
                except subprocess.TimeoutExpired:
                    # The killed process was attached directly to this
                    # terminal (no pipes) -- if it (or a child it spawned,
                    # e.g. aideinit) was in the middle of manipulating
                    # terminal settings when killed, it can leave the tty in
                    # a broken state (no echo, garbled cursor positioning).
                    # Restore it immediately rather than leaving the user
                    # stuck with a terminal that won't accept input.
                    print(f"    [TIMEOUT] Interactive script did not finish "
                          f"within {timeout_s}s -- resetting terminal state.")
                    try:
                        subprocess.run(["stty", "sane"], timeout=5)
                    except Exception:
                        pass
                    exit_code = -1
                    err = f"Interactive script timed out after {timeout_s}s."
                    rule_record["exit_code"] = exit_code
                    rule_record["stdout"] = ""
                    rule_record["stderr"] = err
                except Exception as e:
                    exit_code = -1
                    err = str(e)
                    rule_record["stderr"] = err
            else:
                rule_record["status"] = "permission_denied"
                rule_record["passes"] = passes_log
                counters["permission_denied"] = 1
                return rule_record, counters

        # -- sudoers integrity check --
        if backup_dir is not None:
            valid, vmsg = sudoers_is_valid()
            if not valid:
                print(f"    [SUDOERS BROKEN] {vmsg[:200]}")
                restore_sudoers(backup_dir)
                counters["sudoers_rollback"] = counters.get("sudoers_rollback", 0) + 1
                passes_log.append({"pass": pass_num, "type": "script",
                                    "status": "sudoers_corrupted", "detail": vmsg})
                attempt_history.append({
                    "script": script, "failure_reason": "sudoers_corrupted",
                    "detail": f"This script left /etc/sudoers in an invalid "
                              f"state (visudo -c said: {vmsg}). It has been "
                              f"rolled back. Don't write raw lines into "
                              f"/etc/sudoers or /etc/sudoers.d -- use a "
                              f"properly formatted `Defaults ...` line, or "
                              f"avoid touching sudoers if this rule isn't "
                              f"actually about sudo config.",
                })
                continue

        if exit_code != 0:
            print(f"    [SCRIPT ERROR] exit={exit_code}")
            print(f"    {(err or '')[-150:]}")
            passes_log.append({"pass": pass_num, "type": "script",
                                "status": "script_error", "exit_code": exit_code})
            attempt_history.append({"script": script, "failure_reason": "script_error",
                                     "detail": err})
            continue

        print("    Verifying with oscap...")
        verdict = verify_rule(rule_id)
        rule_record["oscap_result"] = verdict
        passes_log.append({"pass": pass_num, "type": "script", "status": f"oscap_{verdict}"})

        if verdict == "pass":
            print("    PASS")
            rule_record["status"] = "oscap_pass"
            rule_record["passes"] = passes_log
            counters["script_ok"]  = 1
            counters["oscap_pass"] = 1
            return rule_record, counters

        print(f"    {verdict.upper()}")
        attempt_history.append({
            "script": script, "failure_reason": "oscap_fail",
            "detail": f"Script ran fine (exit 0) but oscap re-check still "
                      f"reports '{verdict}' for {rule_id}. The fix didn't "
                      f"actually satisfy the check -- look again at what "
                      f"the rule verifies.",
        })

    # Passes exhausted without a PASS
    rule_record["passes"] = passes_log
    last = attempt_history[-1] if attempt_history else None
    rule_record["status"] = f"unresolved_{last['failure_reason']}" if last else "unresolved"
    counters["passes_exhausted"] = counters.get("passes_exhausted", 0) + 1
    if last and last["failure_reason"] in ("script_error", "timeout", "runtime_error",
                                            "sudoers_corrupted", "repeated_clarify_ignored"):
        counters["script_error"] = counters.get("script_error", 0) + 1
    else:
        counters["oscap_fail"] = counters.get("oscap_fail", 0) + 1
        if rule_record.get("exit_code") == 0:
            counters["script_ok"] = counters.get("script_ok", 0) + 1
    return rule_record, counters


# ─────────────────────────────────────────────────────────────────────────────
# SINGLE MODEL RUN -- iterates rules, skipping anything ground truth says SKIP
# ─────────────────────────────────────────────────────────────────────────────

def run_one_model(model, rule_set, profile_key, auto_approve):
    keep_count = sum(1 for item in rule_set
                      if gt_decision_for_profile(item["gt"], profile_key) == "KEEP")
    skip_count = len(rule_set) - keep_count

    print(f"\n{'='*60}")
    print(f"  MODEL   : {model}")
    print(f"  PROFILE : {profile_key}")
    print(f"  RULES   : {len(rule_set)} total -> {keep_count} KEEP (sent to LLM), "
          f"{skip_count} SKIP (not sent, not remediated)")
    print(f"{'='*60}")

    model_results = {
        "model":   model,
        "profile": profile_key,
        "started": datetime.datetime.now().isoformat(),
        "rules":   {},
        "summary": {
            "attempted": 0, "approved": 0,
            "script_ok": 0, "oscap_pass": 0, "oscap_fail": 0,
            "query_error": 0, "script_error": 0, "rejected": 0,
            "skipped_by_ground_truth": 0,
            "clarifications_asked": 0, "sudoers_rollback": 0,
            "permission_denied": 0, "passes_exhausted": 0,
        },
    }

    for item in rule_set:
        rule_id   = item["rule_id"]
        gt_row    = item["gt"]
        rule_info = item["info"]
        short     = rule_id.replace(PREFIX, "")
        decision  = gt_decision_for_profile(gt_row, profile_key)

        # ---- SKIP handling: never queries the LLM, never touches the system ----
        if decision != "KEEP":
            print(f"\n  [{short}]  gt decision {profile_key}={decision} -- SKIPPING (not sent to LLM)")
            model_results["rules"][rule_id] = {
                "status": "skipped_by_ground_truth",
                "gt_decision": decision,
                "gt_reason": gt_row.get("reason", ""),
            }
            model_results["summary"]["skipped_by_ground_truth"] += 1
            continue

        print(f"\n  [{short}]  (gt decision {profile_key}={decision})")

        rule_record, counters = remediate_rule(
            model=model, rule_id=rule_id, rule_info=rule_info,
            gt_row=gt_row, decision=decision, profile_key=profile_key,
            auto_approve=auto_approve, matched=item["matched"],
        )
        for key, inc in counters.items():
            model_results["summary"][key] = model_results["summary"].get(key, 0) + inc

        model_results["rules"][rule_id] = rule_record
        time.sleep(1)

    s     = model_results["summary"]
    tested = s["attempted"]
    pct   = (s["oscap_pass"] / tested * 100) if tested else 0
    print(f"\n  {model} DONE")
    print(f"  KEEP-tested={tested} Approved={s['approved']} "
          f"PASS={s['oscap_pass']}({pct:.1f}%) FAIL={s['oscap_fail']} "
          f"Err={s['script_error']+s['query_error']} "
          f"Skipped(GT)={s['skipped_by_ground_truth']}")
    print(f"  [Agent loop] Clarifications asked={s.get('clarifications_asked', 0)} "
          f"Sudoers rollbacks={s.get('sudoers_rollback', 0)} "
          f"Permission denied={s.get('permission_denied', 0)} "
          f"Passes exhausted={s.get('passes_exhausted', 0)}")

    model_results["finished"] = datetime.datetime.now().isoformat()
    return model_results


# ─────────────────────────────────────────────────────────────────────────────
# RESET PROMPT BETWEEN MODELS
# ─────────────────────────────────────────────────────────────────────────────

def prompt_reset(prev_model, next_model, break_script, use_snapshot):
    print(f"\n{'='*60}")
    print(f"  {prev_model} complete. Next: {next_model}")
    print(f"{'='*60}")
    if use_snapshot:
        print(f"""
  RESTORE VM SNAPSHOT before continuing.
  Snapshot name: '{SNAPSHOT_NAME}'

  On your HOST machine:
    VBoxManage controlvm "pranjal-garg-VirtualBox" poweroff
    VBoxManage snapshot "pranjal-garg-VirtualBox" restore "{SNAPSHOT_NAME}"
    VBoxManage startvm "pranjal-garg-VirtualBox" --type headless

  Then SSH back in and re-run with --model "{next_model}" --auto
""")
    else:
        print(f"""
  RESET RULES TO FAILING STATE before continuing.
  Run on the VM:
    bash {break_script}

  Then verify rules are failing:
    sudo oscap xccdf eval \\
      --profile xccdf_org.ssgproject.content_profile_cis_level1_workstation \\
      --results agent-test.xml \\
      ~/Downloads/scap-security-guide-0.1.76/ssg-ubuntu2404-ds.xml
""")
    input("  Press Enter when ready for next model...")


# ─────────────────────────────────────────────────────────────────────────────
# SAVE RESULTS
# ─────────────────────────────────────────────────────────────────────────────

def save_results(all_results, profile_key):
    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir  = os.path.join(resolve_path(RESULTS_DIR), f"remediation_{profile_key}_{ts}")
    os.makedirs(run_dir, exist_ok=True)

    for r in all_results:
        fname = r["model"].replace(":", "_").replace("/", "_") + ".json"
        with open(os.path.join(run_dir, fname), "w") as f:
            json.dump(r, f, indent=2)

    lines = [
        "# CIS Remediation Multi-Model Comparison\n\n",
        f"**Profile:** {profile_key}\n\n",
        f"**Timestamp:** {ts}\n\n---\n\n",
        "## Scoreboard (KEEP-category rules only)\n\n",
        "| Model | KEEP Rules Tested | PASS | PASS% | FAIL | Errors | Skipped(GT) |\n",
        "|---|---|---|---|---|---|---|\n",
    ]
    for r in all_results:
        s   = r["summary"]
        tot = s["attempted"]
        pct = (s["oscap_pass"]/tot*100) if tot else 0
        lines.append(f"| {r['model']} | {tot} | {s['oscap_pass']} | "
                     f"{pct:.1f}% | {s['oscap_fail']} | "
                     f"{s['script_error']+s['query_error']} | "
                     f"{s['skipped_by_ground_truth']} |\n")

    lines.append("\n---\n\n## Per-Rule Results (incl. ground truth decision)\n\n")

    all_rids = []
    for r in all_results:
        for rid in r["rules"]:
            if rid not in all_rids:
                all_rids.append(rid)

    hdr = "| Rule | GT Decision | " + " | ".join(r["model"] for r in all_results) + " |\n"
    sep = "|---|---|" + "---|" * len(all_results) + "\n"
    lines += [hdr, sep]

    for rid in all_rids:
        short = rid.replace(PREFIX, "")
        gt_dec = "-"
        for r in all_results:
            rec = r["rules"].get(rid, {})
            if "gt_decision" in rec:
                gt_dec = rec["gt_decision"]
                break
        row = f"| {short} | {gt_dec} |"
        for r in all_results:
            rec    = r["rules"].get(rid, {})
            status = rec.get("status", "-")
            cell   = (
                "PASS" if status == "oscap_pass" else
                "FAIL" if "oscap_fail" in status else
                "SKIP(GT)" if status == "skipped_by_ground_truth" else
                status.replace("_", " ")
            )
            row += f" {cell} |"
        lines.append(row + "\n")

    with open(os.path.join(run_dir, "comparison.md"), "w") as f:
        f.writelines(lines)

    with open(os.path.join(run_dir, "summary.json"), "w") as f:
        json.dump({
            "profile": profile_key, "timestamp": ts,
            "models": [{"model": r["model"], **r["summary"]} for r in all_results],
        }, f, indent=2)

    return run_dir


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    if "--probe" in sys.argv:
        probe_lab_server()
        return

    auto_approve   = "--auto"           in sys.argv
    use_snapshot   = "--snapshot"       in sys.argv
    only_min_delta = "--only-min-delta" in sys.argv
    single_model = None
    cli_profile  = None
    retry_failed_path = None
    only_rules_wanted = None

    if "--model" in sys.argv:
        idx = sys.argv.index("--model")
        if idx + 1 < len(sys.argv):
            single_model = sys.argv[idx + 1]

    if "--retry-failed" in sys.argv:
        idx = sys.argv.index("--retry-failed")
        if idx + 1 < len(sys.argv):
            retry_failed_path = sys.argv[idx + 1]

    if "--only-rules" in sys.argv:
        idx = sys.argv.index("--only-rules")
        if idx + 1 < len(sys.argv):
            only_rules_wanted = {
                r.strip().removeprefix(PREFIX)
                for r in sys.argv[idx + 1].split(",") if r.strip()
            }

    if "--profile" in sys.argv:
        idx = sys.argv.index("--profile")
        if idx + 1 < len(sys.argv):
            cli_profile = sys.argv[idx + 1].strip().upper()
            if cli_profile not in ("MAX", "MIN"):
                print(f"[ERROR] --profile must be MAX or MIN, got '{cli_profile}'")
                sys.exit(1)

    print("\n" + "=" * 60)
    print("  CIS Multi-Model Remediation Pipeline v3.2")
    print(f"  Lab server : {LAB_URL}")
    print(f"  Models     : {', '.join(MODELS)}")
    print(f"  Elevation  : {'sudo bash -c <script>' if ELEVATE_WITH_SUDO else 'bash -c <script> (no sudo)'}")
    print("=" * 60)

    if auto_approve:    print("  [AUTO]     All scripts applied without confirmation.")
    if single_model:    print(f"  [SINGLE]   Running only: {single_model}")
    if use_snapshot:    print(f"  [SNAPSHOT] Will prompt snapshot restore between models.")
    else:               print("  [BREAK]    Will generate break_rules.sh between models.")
    if only_min_delta:  print("  [DELTA]    Only running rules where MIN decision was SKIP "
                               "(assumes KEEP-under-MIN rules already tested).")

    if ELEVATE_WITH_SUDO and auto_approve:
        print("  [NOTE]     --auto + sudo elevation: make sure this user has "
              "passwordless (NOPASSWD) sudo configured, or `sudo` will hang "
              "on a password prompt on the first KEEP rule.")

    if not os.path.exists(resolve_path(GROUND_TRUTH_XLSX)):
        print(f"\n[ERROR] Missing: {GROUND_TRUTH_XLSX}")
        sys.exit(1)

    gt_rows = load_ground_truth(GROUND_TRUTH_XLSX)
    scan_rules = parse_scan_xml(SCAN_RESULT_XML)
    rule_set = build_rule_set(gt_rows, scan_rules)

    profile_key = cli_profile or ask_profile()

    if only_min_delta:
        if profile_key == "MIN":
            print("\n[ERROR] --only-min-delta filters down to rules where MIN "
                  "was SKIP -- running that against profile MIN itself would "
                  "test nothing (every remaining rule is SKIP-under-MIN by "
                  "definition). This flag is meant for --profile MAX runs "
                  "after a full MIN run is already done. Aborting.")
            sys.exit(1)
        rule_set = filter_to_min_skipped(rule_set)

    if only_rules_wanted:
        rule_set = filter_to_only_rules(rule_set, only_rules_wanted)

    if retry_failed_path:
        if not os.path.exists(resolve_path(retry_failed_path)):
            print(f"\n[ERROR] --retry-failed file not found: {retry_failed_path}")
            sys.exit(1)
        rule_set = filter_to_retry_failed(rule_set, retry_failed_path, profile_key)

    keep_n = sum(1 for item in rule_set
                 if gt_decision_for_profile(item["gt"], profile_key) == "KEEP")
    skip_n = len(rule_set) - keep_n
    print(f"\n  Profile {profile_key}: {keep_n} KEEP rules will be sent to the LLM "
          f"and remediated; {skip_n} SKIP rules will be recorded as skipped "
          f"and left untouched.")

    break_script = generate_break_script(rule_set, profile_key)

    models_to_run = MODELS if not single_model else [
        m for m in MODELS if m == single_model
    ]
    if not models_to_run:
        print(f"[ERROR] '{single_model}' not in MODELS list.")
        sys.exit(1)

    os.makedirs(resolve_path(RESULTS_DIR), exist_ok=True)
    all_results = []

    print(f"\n  Running {len(models_to_run)} model(s) sequentially, "
          f"{keep_n} KEEP rule(s) each (SKIP rules recorded, not queried).")
    input("  Press Enter to start with the first model...")

    for i, model in enumerate(models_to_run):
        if i > 0:
            prompt_reset(
                prev_model   = models_to_run[i-1],
                next_model   = model,
                break_script = break_script,
                use_snapshot = use_snapshot,
            )

        result = run_one_model(
            model        = model,
            rule_set     = rule_set,
            profile_key  = profile_key,
            auto_approve = auto_approve,
        )
        all_results.append(result)

        ts    = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = os.path.join(
            resolve_path(RESULTS_DIR),
            f"{model.replace(':','_').replace('/','_')}_{profile_key}_{ts}.json"
        )
        with open(fname, "w") as f:
            json.dump(result, f, indent=2)
        print(f"  Saved: {fname}")

    run_dir = save_results(all_results, profile_key)

    print(f"\n{'='*60}")
    print("  FINAL SCOREBOARD (KEEP-category rules only)")
    print(f"{'='*60}")
    print(f"  {'Model':<25} {'Tested':>7} {'PASS':>6} {'PASS%':>7} {'ERR':>5} {'Skip(GT)':>9}")
    print(f"  {'-'*25} {'------':>7} {'----':>6} {'-----':>7} {'---':>5} {'--------':>9}")
    for r in all_results:
        s   = r["summary"]
        tot = s["attempted"]
        pct = (s["oscap_pass"]/tot*100) if tot else 0
        print(f"  {r['model']:<25} {tot:>7} {s['oscap_pass']:>6} "
              f"{pct:>6.1f}% {s['script_error']+s['query_error']:>5} "
              f"{s['skipped_by_ground_truth']:>9}")

    print(f"\n  Results : {run_dir}")
    print(f"  Table   : {run_dir}/comparison.md")
    print(f"  Reset   : bash {break_script}")


if __name__ == "__main__":
    main()
