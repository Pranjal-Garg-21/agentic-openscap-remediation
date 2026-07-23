#!/usr/bin/env python3
"""
CIS Benchmark Multi-Model Remediation Pipeline v3.1
=====================================================
Fixes vs the version that gave you 13/63 rules + 404s:

  1. GROUND TRUTH FILE BUG (root cause of "only 13 rules"):
     - Your workbook's sheet is named "Ground Truth", not "Ground Truth Grid".
     - Its layout is RULE-PER-ROW (# | Category | Rule | Title | MAX Decision |
       MIN Decision | Conditioned? | Reason) for exactly 63 rules -- NOT the
       role-per-row / rule-per-column grid the old code assumed.
     -> Rewritten load_ground_truth() reads it correctly. All 63 rows load.

  2. ONLY 2 PROFILES (matches your file exactly):
     - MAX = System/Cloud Admin, Public Cloud, Production/Critical
     - MIN = Personal Laptop, Just me, Trusted home network only
     - The old 4-role/follow-up-question system is gone (it didn't match the
       ground truth file at all, which is why matching silently failed down
       to a handful of rows).

  3. ALL 63 RULES ALWAYS RUN, ONE AT A TIME:
     - No KEEP/SKIP filtering. Every rule is attempted for every model.
     - The ground truth KEEP/SKIP + reason for the CHOSEN profile is injected
       into the LLM prompt as context (not used to filter/skip rules) and is
       also recorded in the results JSON so you can compare model behavior
       against the human ground truth afterwards.
     - Rule short-names in the ground truth file don't always exactly match
       the official xccdf rule id in your scan XML (e.g. "hfs_disabled" vs
       "kernel_module_hfs_disabled"). resolve_rule_id() fuzzy-matches them.
       If a rule truly isn't in your scan XML, it still runs -- using the
       ground truth title as the only context -- instead of being skipped.

  4. 404 ON /v1/chat/completions:
     - query_lab() now tries multiple endpoint shapes (OpenAI-style
       /v1/chat/completions, Ollama-native /api/chat) and reports the
       actual response body (not just "404"), so you can see what your lab
       proxy actually expects.
     - Added --probe: hits your lab server with several endpoint/path
       combinations and prints status + body for each, with no rule
       processing, so you can diagnose the correct path in one run.

Usage:
  python3 remediationv3.py --probe                  # diagnose LAB_URL endpoints only
  python3 remediationv3.py                          # interactive, human approval
  python3 remediationv3.py --auto                   # auto-approve all fixes
  python3 remediationv3.py --model qwen2.5:7b       # single model only
  python3 remediationv3.py --profile MAX            # skip the profile prompt
  python3 remediationv3.py --snapshot               # prompt snapshot restore between models
"""

import os, re, sys, json, time, datetime, subprocess, xml.etree.ElementTree as ET
import requests, urllib3
from openpyxl import load_workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

LAB_URL  = os.environ.get("LAB_URL",  "https://10.1.96.96:8443")
LAB_USER = os.environ.get("LAB_USER", "user")
LAB_PASS = os.environ.get("LAB_PASS", "H72j8n19sna")

MODELS = [
    "qwen2.5:7b",
    "gemma2:latest",
    "mistral:latest",
    "granite4.1:8b",
    "gpt-oss:latest",
]

def resolve_path(p):
    if not p or os.path.isabs(p): return p
    if os.path.exists(p): return os.path.abspath(p)
    s_dir = os.path.dirname(os.path.abspath(__file__))
    for base in [s_dir, os.path.join(s_dir, ".."), os.path.join(s_dir, "..", "..")]:
        cand = os.path.abspath(os.path.join(base, p))
        if os.path.exists(cand): return cand
    return os.path.abspath(os.path.join(s_dir, "..", "..", p))

SCAN_RESULT_XML   = "agent-test.xml"
GROUND_TRUTH_XLSX = "CIS_Ground_Truth_FULL_MAX_MIN.xlsx"
GROUND_TRUTH_SHEET_CANDIDATES = ["Ground Truth", "Ground Truth Grid"]
RESULTS_DIR       = "remediation_results"
BENCHMARK_XML     = os.path.expanduser(
    "~/Downloads/scap-security-guide-0.1.76/ssg-ubuntu2404-ds.xml")
SNAPSHOT_NAME     = "baseline-63-rules-broken"

SYSTEM_INFO = {
    "hostname": "pranjal-garg-VirtualBox",
    "kernel":   "6.17.0-23-generic",
    "os":       "Ubuntu 24.04 LTS (Noble Numbat)",
    "arch":     "x86_64",
}

PREFIX = "xccdf_org.ssgproject.content_rule_"

PROFILES = {
    "1": {
        "key": "MAX",
        "label": "MAX -- System/Cloud Admin, Public Cloud, Production/Critical",
    },
    "2": {
        "key": "MIN",
        "label": "MIN -- Personal Laptop, Just me, Trusted home network only",
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# GROUND TRUTH LOADING (rule-per-row layout, 63 rules, MAX/MIN columns)
# ─────────────────────────────────────────────────────────────────────────────

def load_ground_truth(path):
    """
    Returns an ordered list of dicts, one per rule row:
      {num, category, short, title, max_decision, min_decision,
       conditioned, reason}
    Reads the header row dynamically (looks for '#' in col A) instead of
    hardcoding row numbers, so minor formatting shifts don't break it.
    """
    path = resolve_path(path)
    wb = load_workbook(path, data_only=True)

    ws = None
    for name in GROUND_TRUTH_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
        print(f"  [WARN] No sheet named {GROUND_TRUTH_SHEET_CANDIDATES}; "
              f"using first sheet '{ws.title}' instead.")

    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        if row and row[0] == "#" and (row[1] or "").strip().lower().startswith("categ"):
            header_row_idx = i
            break
    if header_row_idx is None:
        raise RuntimeError(
            f"Could not find header row ('#', 'Category', ...) in sheet "
            f"'{ws.title}'. Check the file structure.")

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or row[0] in (None, "", "Totals") or not isinstance(row[0], (int, float)):
            break  # hit the trailing totals/summary block
        num, category, short, title, max_dec, min_dec = row[0:6]
        conditioned = row[6] if len(row) > 6 else None
        reason      = row[7] if len(row) > 7 else None
        if not short:
            continue
        rows.append({
            "num":         int(num),
            "category":    (category or "").strip(),
            "short":       str(short).strip(),
            "title":       (title or "").strip(),
            "max_decision": (max_dec or "").strip().upper(),
            "min_decision": (min_dec or "").strip().upper(),
            "conditioned": (str(conditioned).strip() if conditioned else ""),
            "reason":      (reason or "").strip(),
        })

    print(f"  Ground truth loaded: {len(rows)} rules from sheet '{ws.title}'")
    return rows


def gt_decision_for_profile(gt_row, profile_key):
    if profile_key == "MAX":
        return gt_row["max_decision"] or "KEEP"
    return gt_row["min_decision"] or "KEEP"


# ─────────────────────────────────────────────────────────────────────────────
# SCAN XML PARSING
# ─────────────────────────────────────────────────────────────────────────────

def parse_scan_xml(path):
    path = resolve_path(path)
    rules = {}
    if not os.path.exists(path):
        print(f"  [WARN] Scan XML '{path}' not found -- rules will run with "
              f"ground-truth titles only (no reference fix / description).")
        return rules
    try:
        tree = ET.parse(path)
    except Exception as e:
        print(f"[ERROR] Cannot parse {path}: {e}")
        return rules

    root = tree.getroot()
    for ns_uri in ["http://checklists.nist.gov/xccdf/1.2",
                   "http://checklists.nist.gov/xccdf/1.1"]:
        for rule_el in root.iter(f"{{{ns_uri}}}Rule"):
            rid = rule_el.get("id", "")
            if not rid.startswith(PREFIX):
                continue

            title_el = rule_el.find(f"{{{ns_uri}}}title")
            title    = (title_el.text or "").strip() if title_el is not None else rid

            desc_el  = rule_el.find(f"{{{ns_uri}}}description")
            desc     = "".join(desc_el.itertext()).strip() if desc_el is not None else ""

            severity = rule_el.get("severity", "unknown")

            fixes = {}
            for fix_el in rule_el.findall(f"{{{ns_uri}}}fix"):
                fixes[fix_el.get("system", "fixtext")] = (fix_el.text or "").strip()
            ft = rule_el.find(f"{{{ns_uri}}}fixtext")
            if ft is not None and ft.text:
                fixes["fixtext"] = ft.text.strip()

            fix_text = (
                fixes.get("urn:xccdf:fix:script:sh")
                or fixes.get("fixtext")
                or next(iter(fixes.values()), "")
            )
            fix_system = (
                "sh"      if "urn:xccdf:fix:script:sh" in fixes else
                "fixtext" if "fixtext" in fixes else
                ("other"  if fixes else "none")
            )

            rules[rid] = {
                "title":       title,
                "description": desc[:1200],
                "fix":         fix_text,
                "fix_system":  fix_system,
                "severity":    severity,
            }

    if not rules:
        print(f"\n  No Rule elements found in {path}.")
        print("  Re-run oscap with --results-arf to embed full benchmark content.")
    return rules


# ─────────────────────────────────────────────────────────────────────────────
# GROUND-TRUTH SHORT NAME  <->  OFFICIAL XCCDF RULE ID FUZZY MATCHING
# ─────────────────────────────────────────────────────────────────────────────

def _norm(s):
    return s.replace("-", "_").lower().strip("_")


def resolve_rule_id(gt_short, scan_rules):
    """
    Returns (rule_id, matched_bool).
    matched_bool True  -> rule_id is a real key in scan_rules (full metadata available)
    matched_bool False -> best-guess id (PREFIX + gt_short); rule still runs,
                          just without scan-XML title/description/fix/severity.
    """
    guess = PREFIX + gt_short
    if guess in scan_rules:
        return guess, True

    gt_n = _norm(gt_short)
    for full_id in scan_rules:
        short = full_id[len(PREFIX):] if full_id.startswith(PREFIX) else full_id
        short_n = _norm(short)
        if short_n == gt_n or short_n.endswith("_" + gt_n) or short_n.endswith(gt_n):
            return full_id, True

    # looser fallback: gt short appears anywhere in an official short name
    for full_id in scan_rules:
        short = full_id[len(PREFIX):] if full_id.startswith(PREFIX) else full_id
        if gt_n in _norm(short):
            return full_id, True

    return guess, False


def build_rule_set(gt_rows, scan_rules):
    """
    Produces the ordered list of ALL 63 rules to run, each entry:
      {gt, rule_id, matched, info}
    'info' has title/description/fix/severity -- from scan XML if matched,
    else falls back to the ground truth title only.
    """
    resolved = []
    unmatched = []
    for gt in gt_rows:
        rule_id, matched = resolve_rule_id(gt["short"], scan_rules)
        if matched:
            info = scan_rules[rule_id]
        else:
            unmatched.append(gt["short"])
            info = {
                "title": gt["title"] or gt["short"],
                "description": f"(Not found in scan XML -- CIS category: {gt['category']}. "
                                f"Apply the standard Ubuntu 24.04 remediation for this control.)",
                "fix": "",
                "fix_system": "none",
                "severity": "unknown",
            }
        resolved.append({"gt": gt, "rule_id": rule_id, "matched": matched, "info": info})

    print(f"  Rule set built: {len(resolved)} total "
          f"({len(resolved) - len(unmatched)} matched to scan XML, "
          f"{len(unmatched)} using ground-truth title only)")
    if unmatched:
        print(f"    Unmatched (still WILL run): {unmatched}")
    return resolved


# ─────────────────────────────────────────────────────────────────────────────
# PROFILE SELECTION
# ─────────────────────────────────────────────────────────────────────────────

def ask_profile():
    print("\n" + "=" * 60)
    print("  Select ground-truth profile to test against")
    print("=" * 60)
    for k, v in PROFILES.items():
        print(f"  {k}. {v['label']}")
    while True:
        c = input("\nProfile number: ").strip()
        if c in PROFILES:
            return PROFILES[c]["key"]
        print("Invalid, try again.")


# ─────────────────────────────────────────────────────────────────────────────
# LAB SERVER QUERY  (with multi-endpoint fallback + diagnostics)
# ─────────────────────────────────────────────────────────────────────────────

def _try_openai_chat(model, prompt, max_tokens, timeout):
    url = f"{LAB_URL}/v1/chat/completions"
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
        "temperature": 0.2,
        "stream": False,
    }
    resp = requests.post(url, json=payload, auth=(LAB_USER, LAB_PASS),
                          verify=False, timeout=timeout)
    return resp, url


def _try_ollama_chat(model, prompt, max_tokens, timeout):
    url = f"{LAB_URL}/api/chat"
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "stream": False,
        "options": {"temperature": 0.2, "num_predict": max_tokens},
    }
    resp = requests.post(url, json=payload, auth=(LAB_USER, LAB_PASS),
                          verify=False, timeout=timeout)
    return resp, url


def _try_ollama_generate(model, prompt, max_tokens, timeout):
    url = f"{LAB_URL}/api/generate"
    payload = {
        "model": model,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": 0.2, "num_predict": max_tokens},
    }
    resp = requests.post(url, json=payload, auth=(LAB_USER, LAB_PASS),
                          verify=False, timeout=timeout)
    return resp, url


_ENDPOINT_TRIED_ORDER = [_try_openai_chat, _try_ollama_chat, _try_ollama_generate]
_WORKING_ENDPOINT = [None]  # cache which shape worked, so we don't re-probe every call


def query_lab(model, prompt, max_tokens=900, timeout=120):
    attempts = []

    fns = _ENDPOINT_TRIED_ORDER
    if _WORKING_ENDPOINT[0] is not None:
        fns = [_WORKING_ENDPOINT[0]] + [f for f in _ENDPOINT_TRIED_ORDER if f is not _WORKING_ENDPOINT[0]]

    for fn in fns:
        try:
            resp, url = fn(model, prompt, max_tokens, timeout)
        except requests.exceptions.RequestException as e:
            attempts.append(f"{fn.__name__} -> connection error: {e}")
            continue

        if resp.status_code == 404:
            attempts.append(f"{fn.__name__} [{url}] -> 404: {resp.text[:200]}")
            continue
        if resp.status_code == 401 or resp.status_code == 403:
            raise RuntimeError(
                f"Auth rejected ({resp.status_code}) at {url}. "
                f"Check LAB_USER/LAB_PASS. Body: {resp.text[:200]}")
        if not resp.ok:
            attempts.append(f"{fn.__name__} [{url}] -> {resp.status_code}: {resp.text[:200]}")
            continue

        try:
            data = resp.json()
        except ValueError:
            attempts.append(f"{fn.__name__} [{url}] -> non-JSON body: {resp.text[:200]}")
            continue

        _WORKING_ENDPOINT[0] = fn  # remember what worked for next call

        if "choices" in data:
            return data["choices"][0]["message"]["content"].strip()
        elif "message" in data and isinstance(data["message"], dict):
            return data["message"]["content"].strip()
        elif "response" in data:  # ollama /api/generate shape
            return data["response"].strip()
        else:
            attempts.append(f"{fn.__name__} [{url}] -> unrecognized JSON keys: {list(data.keys())}")
            continue

    raise RuntimeError(
        "All endpoint shapes failed for model '" + model + "':\n  " +
        "\n  ".join(attempts)
    )


def probe_lab_server():
    """--probe mode: hit several endpoints/paths and print raw results so you
    can see exactly what your lab server exposes."""
    print("\n" + "=" * 60)
    print(f"  PROBING LAB SERVER: {LAB_URL}")
    print("=" * 60)

    checks = [
        ("GET",  "/v1/models"),
        ("GET",  "/api/tags"),
        ("GET",  "/"),
        ("GET",  "/health"),
    ]
    for method, path in checks:
        url = LAB_URL + path
        try:
            resp = requests.request(method, url, auth=(LAB_USER, LAB_PASS),
                                     verify=False, timeout=15)
            print(f"  {method} {path:20s} -> {resp.status_code}  {resp.text[:150]!r}")
        except requests.exceptions.RequestException as e:
            print(f"  {method} {path:20s} -> ERROR: {e}")

    test_model = MODELS[0]
    print(f"\n  Now trying actual chat-completion shapes with model '{test_model}':")
    for fn in _ENDPOINT_TRIED_ORDER:
        try:
            resp, url = fn(test_model, "Say OK.", 20, 30)
            print(f"  {fn.__name__:20s} [{url}] -> {resp.status_code}  {resp.text[:200]!r}")
        except requests.exceptions.RequestException as e:
            print(f"  {fn.__name__:20s} -> ERROR: {e}")

    print("\n  Whichever shape/path returned 200 above with a real completion is")
    print("  the one to keep; tell me the working path and I'll hardcode it so")
    print("  every call skips straight to it instead of probing.")


def strip_code_fences(text):
    text = re.sub(r"^```[a-zA-Z]*\n?", "", text.strip())
    text = re.sub(r"\n?```$", "", text)
    return text.strip()


# ─────────────────────────────────────────────────────────────────────────────
# PROMPT BUILDER  (now includes ground-truth decision + reason as context)
# ─────────────────────────────────────────────────────────────────────────────

def build_prompt(rule_id, rule_info, profile_key, gt_row):
    ephemeral  = False  # no more dev-stack/downtime follow-ups; always live bash
    style_note = "Give the fix as a runnable bash script (it will be run directly with `bash -c`)."

    gt_decision = gt_decision_for_profile(gt_row, profile_key)
    gt_reason   = gt_row.get("reason") or "(no reason recorded)"

    profile_desc = (
        "System/Cloud Administrator running a production, internet-facing "
        "workload in a public cloud"
        if profile_key == "MAX" else
        "individual on a personal laptop, used only by them, on a trusted "
        "home network"
    )

    return f"""You are a Linux system hardening expert. Write a remediation script for
the following CIS/OpenSCAP rule on this specific host.

Host: {SYSTEM_INFO['hostname']} | {SYSTEM_INFO['os']} | kernel {SYSTEM_INFO['kernel']} | {SYSTEM_INFO['arch']}
Profile: {profile_key} -- {profile_desc}

Rule ID:     {rule_id}
Title:       {rule_info['title']}
Severity:    {rule_info['severity']}
Description: {rule_info['description']}

Ground-truth expert assessment for this profile:
  Decision: {gt_decision}
  Reason:   {gt_reason}
(This is provided as context on why the control matters or doesn't for this
profile. Still write the correct technical remediation script regardless of
whether the decision is KEEP or SKIP -- we are evaluating remediation
correctness, not asking you to decide whether to apply it.)

Reference fix from benchmark (format: {rule_info.get('fix_system', 'none')}):
{rule_info['fix'] or '(none provided -- write the standard Ubuntu 24.04 remediation yourself)'}

If the reference fix is bash (sh), adapt it directly.
If it is Ansible/Puppet/blueprint or missing, translate the intent into plain bash.

{style_note}
Output ONLY the script -- no prose, no markdown fences, no explanation.
"""


# ─────────────────────────────────────────────────────────────────────────────
# OSCAP SINGLE-RULE VERIFICATION
# ─────────────────────────────────────────────────────────────────────────────

def verify_rule(rule_id):
    cmd = (
        f"sudo oscap xccdf eval --rule {rule_id} "
        f"{BENCHMARK_XML} 2>/dev/null | grep 'Result'"
    )
    try:
        out = subprocess.check_output(
            cmd, shell=True, text=True, timeout=90
        ).strip().lower()
        if "pass" in out:            return "pass"
        elif "fail" in out:          return "fail"
        elif "notapplicable" in out: return "notapplicable"
        else:                        return "notchecked"
    except subprocess.TimeoutExpired:
        return "timeout"
    except Exception:
        return "error"


# ─────────────────────────────────────────────────────────────────────────────
# BREAK-RULES SCRIPT GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

BREAK_COMMANDS = {
    "accounts_passwords_pam_faillock_deny":
        "sudo sed -i '/faillock/d' /etc/pam.d/common-auth /etc/pam.d/common-account 2>/dev/null || true",
    "accounts_passwords_pam_faillock_enabled":
        "sudo sed -i '/pam_faillock/d' /etc/pam.d/common-auth /etc/pam.d/common-account 2>/dev/null || true",
    "accounts_passwords_pam_faillock_unlock_time":
        "sudo sed -i '/faillock/d' /etc/pam.d/common-auth 2>/dev/null || true",
    "accounts_password_pam_dcredit":
        "sudo sed -i '/dcredit/d' /etc/security/pwquality.conf 2>/dev/null || true",
    "accounts_password_pam_minlen":
        "sudo sed -i '/minlen/d' /etc/security/pwquality.conf 2>/dev/null || true",
    "accounts_password_pam_ucredit":
        "sudo sed -i '/ucredit/d' /etc/security/pwquality.conf 2>/dev/null || true",
    "accounts_password_pam_unix_no_remember":
        "sudo sed -i '/remember/d' /etc/pam.d/common-password 2>/dev/null || true",
    "accounts_password_pam_unix_authtok":
        "sudo sed -i 's/ *sha512//' /etc/pam.d/common-password 2>/dev/null || true",
    "set_password_hashing_algorithm_systemauth":
        "sudo sed -i 's/ *sha512//' /etc/pam.d/common-password 2>/dev/null || true",
    "no_empty_passwords_unix":
        "sudo sed -i 's/ nullok_secure//g' /etc/pam.d/common-auth 2>/dev/null || true",
    "accounts_tmout":
        "sudo sed -i '/TMOUT/d' /etc/bash.bashrc /etc/profile /etc/profile.d/*.sh 2>/dev/null || true",
    "accounts_umask_etc_bashrc":
        "sudo sed -i '/umask 027/d; /umask 077/d' /etc/bash.bashrc 2>/dev/null || true",
    "umask_etc_login_defs":
        "sudo sed -i 's/^UMASK.*/UMASK 022/' /etc/login.defs 2>/dev/null || true",
    "sudo_custom_logfile":
        "sudo sed -i '/logfile/d' /etc/sudoers 2>/dev/null; sudo rm -f /etc/sudoers.d/logfile 2>/dev/null || true",
    "sudo_require_reauthentication":
        "sudo rm -f /etc/sudoers.d/timeout 2>/dev/null; echo 'Defaults timestamp_timeout=15' | sudo tee /etc/sudoers.d/noauth > /dev/null",
    "sudo_remove_no_authenticate":
        "echo 'Defaults !authenticate' | sudo tee /etc/sudoers.d/zz_break_authenticate > /dev/null",
    "package_apparmor-utils_installed":
        "sudo apt-get remove -y apparmor-utils 2>/dev/null || true",
    "grub2_enable_apparmor":
        "sudo sed -i 's/ apparmor=1 security=apparmor//' /etc/default/grub 2>/dev/null; sudo update-grub 2>/dev/null || true",
    "all_apparmor_profiles_in_enforce_complain_mode":
        "sudo aa-complain /etc/apparmor.d/* 2>/dev/null || true",
    "sysctl_net_ipv6_conf_all_forwarding":
        "sudo sed -i '/net.ipv6.conf.all.forwarding/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv6.conf.all.forwarding=1 2>/dev/null || true",
    "sysctl_net_ipv4_conf_all_accept_redirects":
        "sudo sed -i '/accept_redirects/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.conf.all.accept_redirects=1 2>/dev/null || true",
    "sysctl_net_ipv4_conf_all_log_martians":
        "sudo sed -i '/log_martians/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.conf.all.log_martians=0 2>/dev/null || true",
    "sysctl_net_ipv4_conf_all_rp_filter":
        "sudo sed -i '/rp_filter/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.conf.all.rp_filter=0 2>/dev/null || true",
    "sysctl_net_ipv4_tcp_syncookies":
        "sudo sed -i '/syncookies/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.tcp_syncookies=0 2>/dev/null || true",
    "sysctl_net_ipv4_conf_all_send_redirects":
        "sudo sed -i '/send_redirects/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.conf.all.send_redirects=1 2>/dev/null || true",
    "sysctl_net_ipv4_ip_forward":
        "sudo sed -i '/ip_forward/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.ip_forward=1 2>/dev/null || true",
    "sysctl_kernel_randomize_va_space":
        "sudo sed -i '/randomize_va_space/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w kernel.randomize_va_space=0 2>/dev/null || true",
    "sysctl_fs_suid_dumpable":
        "sudo sed -i '/suid_dumpable/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w fs.suid_dumpable=1 2>/dev/null || true",
    "disable_users_coredumps":
        "sudo sed -i '/hard core/d; /soft core/d' /etc/security/limits.conf 2>/dev/null; echo '* soft core unlimited' | sudo tee -a /etc/security/limits.conf",
    "kernel_module_cramfs_disabled":
        "sudo rm -f /etc/modprobe.d/cramfs.conf 2>/dev/null || true",
    "kernel_module_hfs_disabled":
        "sudo rm -f /etc/modprobe.d/hfs.conf 2>/dev/null || true",
    "kernel_module_hfsplus_disabled":
        "sudo rm -f /etc/modprobe.d/hfsplus.conf 2>/dev/null || true",
    "kernel_module_jffs2_disabled":
        "sudo rm -f /etc/modprobe.d/jffs2.conf 2>/dev/null || true",
    "mount_option_dev_shm_nodev":
        "sudo sed -i '/\\/dev\\/shm/d' /etc/fstab; sudo mount -o remount /dev/shm 2>/dev/null || true",
    "mount_option_dev_shm_noexec":
        "sudo sed -i '/\\/dev\\/shm/d' /etc/fstab; sudo mount -o remount /dev/shm 2>/dev/null || true",
    "mount_option_dev_shm_nosuid":
        "sudo sed -i '/\\/dev\\/shm/d' /etc/fstab; sudo mount -o remount /dev/shm 2>/dev/null || true",
    "file_permissions_crontab":
        "sudo chmod 0644 /etc/crontab 2>/dev/null || true",
    "file_permissions_cron_allow":
        "sudo chmod 0644 /etc/cron.allow 2>/dev/null || true",
    "file_permissions_cron_d":
        "sudo chmod 0755 /etc/cron.d 2>/dev/null || true",
    "file_permissions_cron_daily":
        "sudo chmod 0755 /etc/cron.daily 2>/dev/null || true",
    "file_owner_cron_allow":
        "sudo chown nobody:nogroup /etc/cron.allow 2>/dev/null || true",
    "file_groupowner_cron_allow":
        "sudo chgrp root /etc/cron.allow 2>/dev/null || true",
    "file_groupowner_backup_etc_gshadow":
        "sudo chgrp root /etc/gshadow- 2>/dev/null || true",
    "package_ftp_removed":
        "sudo apt-get install -y ftp 2>/dev/null || true",
    "package_tnftp_removed":
        "sudo apt-get install -y tnftp 2>/dev/null || true",
    "package_openldap-clients_removed":
        "sudo apt-get install -y ldap-utils 2>/dev/null || true",
    "package_rsync_removed":
        "sudo apt-get install -y rsync 2>/dev/null || true",
    "service_rsyncd_disabled":
        "sudo systemctl enable rsync 2>/dev/null || sudo systemctl enable rsyncd 2>/dev/null || true",
    "package_telnet_removed":
        "sudo apt-get install -y telnet 2>/dev/null || true",
    "package_vsftpd_removed":
        "sudo apt-get install -y vsftpd 2>/dev/null || true",
    "service_vsftpd_disabled":
        "sudo systemctl enable vsftpd 2>/dev/null || true",
    "package_nis_removed":
        "sudo apt-get install -y nis 2>/dev/null || true",
    "package_rpcbind_removed":
        "sudo apt-get install -y rpcbind 2>/dev/null || true",
    "package_ypserv_removed":
        "sudo apt-get install -y nis 2>/dev/null || true",
}


def generate_break_script(rule_set, output_path="break_rules.sh"):
    not_found = []
    lines = [
        "#!/bin/bash",
        "# Auto-generated by remediationv3.py",
        "# Resets rules back to FAILING state between models",
        "# Run between models: bash break_rules.sh",
        "echo '=== Resetting rules to failing state ==='",
        "",
    ]
    for item in rule_set:
        short = item["rule_id"].replace(PREFIX, "")
        gt_short = item["gt"]["short"]
        cmd = BREAK_COMMANDS.get(short) or BREAK_COMMANDS.get(gt_short)
        if cmd:
            lines.append(f"echo '  Breaking: {short}'")
            lines.append(cmd)
        else:
            not_found.append(short)

    lines += ["", "echo '=== Reset complete ==='"]
    if not_found:
        lines.append(f"echo 'No break command for: {not_found}'")

    with open(output_path, "w") as f:
        f.write("\n".join(lines) + "\n")
    os.chmod(output_path, 0o755)

    print(f"\n  Break script written to: {output_path}")
    if not_found:
        print(f"  Rules without break command (add manually): {not_found}")
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# SINGLE MODEL RUN -- iterates ALL 63 rules, one query at a time
# ─────────────────────────────────────────────────────────────────────────────

def run_one_model(model, rule_set, profile_key, auto_approve):
    print(f"\n{'='*60}")
    print(f"  MODEL   : {model}")
    print(f"  PROFILE : {profile_key}")
    print(f"  RULES   : {len(rule_set)}")
    print(f"{'='*60}")

    model_results = {
        "model":   model,
        "profile": profile_key,
        "started": datetime.datetime.now().isoformat(),
        "rules":   {},
        "summary": {
            "attempted": 0, "approved": 0,
            "script_ok": 0, "oscap_pass": 0, "oscap_fail": 0,
            "query_error": 0, "script_error": 0, "rejected": 0,
        },
    }

    for item in rule_set:
        rule_id   = item["rule_id"]
        gt_row    = item["gt"]
        rule_info = item["info"]
        short     = rule_id.replace(PREFIX, "")

        print(f"\n  [{short}]  (gt decision {profile_key}="
              f"{gt_decision_for_profile(gt_row, profile_key)})")

        # Query lab model
        print(f"    Querying {model}...")
        try:
            raw = query_lab(model, build_prompt(rule_id, rule_info, profile_key, gt_row))
        except Exception as e:
            print(f"    [QUERY ERROR] {e}")
            model_results["rules"][rule_id] = {
                "status": "query_error", "error": str(e),
                "gt_decision": gt_decision_for_profile(gt_row, profile_key),
            }
            model_results["summary"]["query_error"] += 1
            time.sleep(2)
            continue

        script = strip_code_fences(raw)
        model_results["summary"]["attempted"] += 1

        print("\n    Proposed fix:")
        print("    " + "-"*52)
        for line in script.split("\n")[:20]:
            print(f"    {line}")
        if script.count("\n") > 20:
            print(f"    ... ({script.count(chr(10))-20} more lines)")
        print("    " + "-"*52)

        if auto_approve:
            approved = True
            print("    [AUTO] Applying.")
        else:
            ans = input("    Apply? [y/n/s=show full]: ").strip().lower()
            if ans == "s":
                print(f"\n{script}\n")
                ans = input("    Apply? [y/n]: ").strip().lower()
            approved = (ans == "y")

        rule_record = {
            "script": script,
            "approved": approved,
            "matched_in_scan": item["matched"],
            "gt_decision": gt_decision_for_profile(gt_row, profile_key),
            "gt_reason": gt_row.get("reason", ""),
        }

        if not approved:
            rule_record["status"] = "rejected"
            model_results["summary"]["rejected"] += 1
            model_results["rules"][rule_id] = rule_record
            continue

        model_results["summary"]["approved"] += 1

        try:
            proc = subprocess.run(
                ["bash", "-c", script],
                capture_output=True, text=True, timeout=120,
            )
            rule_record["exit_code"] = proc.returncode
            rule_record["stdout"]    = proc.stdout[-2000:]
            rule_record["stderr"]    = proc.stderr[-2000:]

            if proc.returncode != 0:
                print(f"    [SCRIPT ERROR] exit={proc.returncode}")
                print(f"    {proc.stderr[-150:]}")
                rule_record["status"] = "script_error"
                model_results["summary"]["script_error"] += 1
            else:
                model_results["summary"]["script_ok"] += 1
                print("    Verifying with oscap...")
                verdict = verify_rule(rule_id)
                rule_record["oscap_result"] = verdict
                if verdict == "pass":
                    print("    PASS")
                    rule_record["status"] = "oscap_pass"
                    model_results["summary"]["oscap_pass"] += 1
                else:
                    print(f"    {verdict.upper()}")
                    rule_record["status"] = f"oscap_{verdict}"
                    model_results["summary"]["oscap_fail"] += 1

        except subprocess.TimeoutExpired:
            print("    [TIMEOUT]")
            rule_record["status"] = "timeout"
            model_results["summary"]["script_error"] += 1
        except Exception as e:
            rule_record["status"] = "error"
            rule_record["error"]  = str(e)
            model_results["summary"]["script_error"] += 1

        model_results["rules"][rule_id] = rule_record
        time.sleep(1)

    s     = model_results["summary"]
    total = len(rule_set)
    pct   = (s["oscap_pass"] / total * 100) if total else 0
    print(f"\n  {model} DONE")
    print(f"  Attempted={s['attempted']} Approved={s['approved']} "
          f"PASS={s['oscap_pass']}({pct:.1f}%) FAIL={s['oscap_fail']} "
          f"Err={s['script_error']+s['query_error']}")

    model_results["finished"] = datetime.datetime.now().isoformat()
    return model_results


# ─────────────────────────────────────────────────────────────────────────────
# RESET PROMPT BETWEEN MODELS
# ─────────────────────────────────────────────────────────────────────────────

def prompt_reset(prev_model, next_model, break_script, use_snapshot):
    print(f"\n{'='*60}")
    print(f"  {prev_model} complete. Next: {next_model}")
    print(f"{'='*60}")
    if use_snapshot:
        print(f"""
  RESTORE VM SNAPSHOT before continuing.
  Snapshot name: '{SNAPSHOT_NAME}'

  On your HOST machine:
    VBoxManage controlvm "pranjal-garg-VirtualBox" poweroff
    VBoxManage snapshot "pranjal-garg-VirtualBox" restore "{SNAPSHOT_NAME}"
    VBoxManage startvm "pranjal-garg-VirtualBox" --type headless

  Then SSH back in and re-run with --model "{next_model}" --auto
""")
    else:
        print(f"""
  RESET RULES TO FAILING STATE before continuing.
  Run on the VM:
    bash {break_script}

  Then verify rules are failing:
    sudo oscap xccdf eval \\
      --profile xccdf_org.ssgproject.content_profile_cis_level1_workstation \\
      --results agent-test.xml \\
      ~/Downloads/scap-security-guide-0.1.76/ssg-ubuntu2404-ds.xml
""")
    input("  Press Enter when ready for next model...")


# ─────────────────────────────────────────────────────────────────────────────
# SAVE RESULTS
# ─────────────────────────────────────────────────────────────────────────────

def save_results(all_results, profile_key):
    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir  = os.path.join(resolve_path(RESULTS_DIR), f"remediation_{profile_key}_{ts}")
    os.makedirs(run_dir, exist_ok=True)

    for r in all_results:
        fname = r["model"].replace(":", "_").replace("/", "_") + ".json"
        with open(os.path.join(run_dir, fname), "w") as f:
            json.dump(r, f, indent=2)

    lines = [
        "# CIS Remediation Multi-Model Comparison\n\n",
        f"**Profile:** {profile_key}\n\n",
        f"**Timestamp:** {ts}\n\n---\n\n",
        "## Scoreboard\n\n",
        "| Model | Rules | PASS | PASS% | FAIL | Errors |\n",
        "|---|---|---|---|---|---|\n",
    ]
    for r in all_results:
        s   = r["summary"]
        tot = len(r["rules"])
        pct = (s["oscap_pass"]/tot*100) if tot else 0
        lines.append(f"| {r['model']} | {tot} | {s['oscap_pass']} | "
                     f"{pct:.1f}% | {s['oscap_fail']} | "
                     f"{s['script_error']+s['query_error']} |\n")

    lines.append("\n---\n\n## Per-Rule Results (incl. ground truth decision)\n\n")

    all_rids = []
    for r in all_results:
        for rid in r["rules"]:
            if rid not in all_rids:
                all_rids.append(rid)

    hdr = "| Rule | GT Decision | " + " | ".join(r["model"] for r in all_results) + " |\n"
    sep = "|---|---|" + "---|" * len(all_results) + "\n"
    lines += [hdr, sep]

    for rid in all_rids:
        short = rid.replace(PREFIX, "")
        gt_dec = "-"
        for r in all_results:
            rec = r["rules"].get(rid, {})
            if "gt_decision" in rec:
                gt_dec = rec["gt_decision"]
                break
        row = f"| {short} | {gt_dec} |"
        for r in all_results:
            rec    = r["rules"].get(rid, {})
            status = rec.get("status", "-")
            cell   = (
                "PASS" if status == "oscap_pass" else
                "FAIL" if "oscap_fail" in status else
                status.replace("_", " ")
            )
            row += f" {cell} |"
        lines.append(row + "\n")

    with open(os.path.join(run_dir, "comparison.md"), "w") as f:
        f.writelines(lines)

    with open(os.path.join(run_dir, "summary.json"), "w") as f:
        json.dump({
            "profile": profile_key, "timestamp": ts,
            "models": [{"model": r["model"], **r["summary"]} for r in all_results],
        }, f, indent=2)

    return run_dir


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    if "--probe" in sys.argv:
        probe_lab_server()
        return

    auto_approve = "--auto"     in sys.argv
    use_snapshot = "--snapshot" in sys.argv
    single_model = None
    cli_profile  = None

    if "--model" in sys.argv:
        idx = sys.argv.index("--model")
        if idx + 1 < len(sys.argv):
            single_model = sys.argv[idx + 1]

    if "--profile" in sys.argv:
        idx = sys.argv.index("--profile")
        if idx + 1 < len(sys.argv):
            cli_profile = sys.argv[idx + 1].strip().upper()
            if cli_profile not in ("MAX", "MIN"):
                print(f"[ERROR] --profile must be MAX or MIN, got '{cli_profile}'")
                sys.exit(1)

    print("\n" + "=" * 60)
    print("  CIS Multi-Model Remediation Pipeline v3.1")
    print(f"  Lab server : {LAB_URL}")
    print(f"  Models     : {', '.join(MODELS)}")
    print("=" * 60)

    if auto_approve:  print("  [AUTO]     All scripts applied without confirmation.")
    if single_model:  print(f"  [SINGLE]   Running only: {single_model}")
    if use_snapshot:  print(f"  [SNAPSHOT] Will prompt snapshot restore between models.")
    else:             print("  [BREAK]    Will generate break_rules.sh between models.")

    if not os.path.exists(resolve_path(GROUND_TRUTH_XLSX)):
        print(f"\n[ERROR] Missing: {GROUND_TRUTH_XLSX}")
        sys.exit(1)

    gt_rows = load_ground_truth(GROUND_TRUTH_XLSX)
    scan_rules = parse_scan_xml(SCAN_RESULT_XML)
    rule_set = build_rule_set(gt_rows, scan_rules)

    profile_key = cli_profile or ask_profile()
    print(f"\n  Running ALL {len(rule_set)} rules for profile: {profile_key}")

    break_script = generate_break_script(rule_set)

    models_to_run = MODELS if not single_model else [
        m for m in MODELS if m == single_model
    ]
    if not models_to_run:
        print(f"[ERROR] '{single_model}' not in MODELS list.")
        sys.exit(1)

    os.makedirs(resolve_path(RESULTS_DIR), exist_ok=True)
    all_results = []

    print(f"\n  Running {len(models_to_run)} model(s) sequentially, "
          f"{len(rule_set)} rules each, one rule per request.")
    input("  Press Enter to start with the first model...")

    for i, model in enumerate(models_to_run):
        if i > 0:
            prompt_reset(
                prev_model   = models_to_run[i-1],
                next_model   = model,
                break_script = break_script,
                use_snapshot = use_snapshot,
            )

        result = run_one_model(
            model        = model,
            rule_set     = rule_set,
            profile_key  = profile_key,
            auto_approve = auto_approve,
        )
        all_results.append(result)

        ts    = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = os.path.join(
            resolve_path(RESULTS_DIR),
            f"{model.replace(':','_').replace('/','_')}_{profile_key}_{ts}.json"
        )
        with open(fname, "w") as f:
            json.dump(result, f, indent=2)
        print(f"  Saved: {fname}")

    run_dir = save_results(all_results, profile_key)

    print(f"\n{'='*60}")
    print("  FINAL SCOREBOARD")
    print(f"{'='*60}")
    print(f"  {'Model':<25} {'Rules':>6} {'PASS':>6} {'PASS%':>7} {'ERR':>5}")
    print(f"  {'-'*25} {'----':>6} {'----':>6} {'-----':>7} {'---':>5}")
    for r in all_results:
        s   = r["summary"]
        tot = len(r["rules"])
        pct = (s["oscap_pass"]/tot*100) if tot else 0
        print(f"  {r['model']:<25} {tot:>6} {s['oscap_pass']:>6} "
              f"{pct:>6.1f}% {s['script_error']+s['query_error']:>5}")

    print(f"\n  Results : {run_dir}")
    print(f"  Table   : {run_dir}/comparison.md")
    print(f"  Reset   : bash {break_script}")


if __name__ == "__main__":
    main()