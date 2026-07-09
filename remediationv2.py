#!/usr/bin/env python3
"""
CIS Benchmark Multi-Model Remediation Pipeline
================================================
Takes the role/profile answers (same flow as analysisv2.py), looks up the
KEEP rules for that combo from CIS_Ground_Truth_FINAL.xlsx, then for each
KEEP rule asks each of the 4 candidate models to write a remediation
script from the rule's description + fix text pulled out of agent-test.xml.
Every proposed script is shown to you and only runs after an explicit
y/n confirmation. Results are logged per model so you can compare who
actually fixes what.
"""

import os
import re
import sys
import json
import time
import datetime
import subprocess
import xml.etree.ElementTree as ET

import requests
from openpyxl import load_workbook

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

# Set this as an env var instead of hardcoding it in the file — the version
# in analysisv2.py has a live key committed in plaintext, which should be
# rotated on build.nvidia.com regardless of what happens with this script.
NVIDIA_API_KEY = os.environ.get("NVIDIA_API_KEY", "")
NVIDIA_URL = "https://integrate.api.nvidia.com/v1/chat/completions"

MODELS = [
    "google/gemma-4-31b-it",
    "mistralai/mistral-large-3-675b-instruct-2512",
    "moonshotai/kimi-k2.6",
    "openai/gpt-oss-120b",
]

SCAN_RESULT_XML = "agent-test.xml"
GROUND_TRUTH_XLSX = "CIS_Ground_Truth_FINAL.xlsx"
RESULTS_DIR = "remediation_results"

SYSTEM_INFO = {
    "hostname": "pranjal-garg-IdeaPad-Slim-5-14IRL8",
    "kernel": "6.17.0-29-generic #29~24.04.1-Ubuntu SMP PREEMPT_DYNAMIC Mon May 11 10:30:58 UTC 2",
    "os": "Ubuntu 24.04 LTS (Noble Numbat)",
    "arch": "x86_64",
}

PREFIX = "xccdf_org.ssgproject.content_rule_"

# ─────────────────────────────────────────────
# ROLE / FOLLOW-UP QUESTIONS (mirrors analysisv2.py)
# ─────────────────────────────────────────────

ROLES = {
    "1": "Personal Laptop / Home User",
    "2": "Student / Security Learner / Researcher",
    "3": "Software Developer",
    "4": "System / Cloud Administrator",
}

FOLLOWUP_QUESTIONS = {
    "Personal Laptop / Home User": [
        {"q": "Who physically uses this computer?",
         "options": {"1": "Just me (Low risk of physical tampering)",
                     "2": "Shared with family or roommates (Moderate risk, needs basic user isolation)"},
         "key": "physical_access", "multi": False},
        {"q": "Where do you connect?",
         "options": {"1": "Only trusted home/private networks (Standard firewall is fine)",
                     "2": "Frequently on public campus or cafe Wi-Fi (Needs aggressive network hardening)"},
         "key": "network_environment", "multi": False},
    ],
    "Student / Security Learner / Researcher": [
        {"q": "What do you actually use this computer for? (Select ALL that apply)",
         "options": {"1": "Coding & Development (Writing code, running local web servers, or building apps)",
                     "2": "Security & Hacking (Playing with network scanners, testing vulnerabilities, or CTFs)",
                     "3": "General Technical Work (Basic scripting, data analysis, and standard terminal usage)"},
         "key": "learning_workloads", "multi": True},
        {"q": "How comfortable are you with the Linux terminal?",
         "options": {"1": "Beginner (Explain exactly what the commands do before I run them)",
                     "2": "Advanced (Just give me the raw commands or config file edits, I know what they do)"},
         "key": "technical_depth", "multi": False},
    ],
    "Software Developer": [
        {"q": "What are you building? (Select ALL that apply)",
         "options": {"1": "Web / Full-Stack (MERN, React Native, Node.js - needs local port access)",
                     "2": "Systems / Low-Level (C/C++, Linux kernels - needs deep system execution rights)",
                     "3": "Containerized Apps (Docker/Podman - relies on virtual networking)"},
         "key": "dev_stack", "multi": True},
        {"q": "Does this machine accept external connections?",
         "options": {"1": "Yes, I run local servers/APIs that teammates or external tools connect to",
                     "2": "No, it's strictly offline compiling and local-only testing"},
         "key": "network_exposure", "multi": False},
    ],
    "System / Cloud Administrator": [
        {"q": "How sensitive is this server to downtime?",
         "options": {"1": "Production / Critical",
                     "2": "Internal / Workstation",
                     "3": "Ephemeral (config/Dockerfile fixes only, no live bash)"},
         "key": "downtime_sensitivity", "multi": False},
        {"q": "Where does this infrastructure live?",
         "options": {"1": "Public Cloud",
                     "2": "Internal Corporate Network",
                     "3": "Local Virtual Machine (Sandboxed environment)"},
         "key": "infrastructure_location", "multi": False},
    ],
}

# ─────────────────────────────────────────────
# CLI — ROLE + FOLLOW-UP QUESTIONS
# ─────────────────────────────────────────────

def ask_role():
    print("\n" + "=" * 55)
    print("  CIS Remediation Pipeline — select your role")
    print("=" * 55)
    for k, v in ROLES.items():
        print(f"  {k}. {v}")
    while True:
        choice = input("\nRole number: ").strip()
        if choice in ROLES:
            return ROLES[choice]
        print("Invalid choice, try again.")


def ask_followups(role):
    questions = FOLLOWUP_QUESTIONS.get(role, [])
    profile = {}
    for q in questions:
        print(f"\n{q['q']}")
        for k, v in q["options"].items():
            print(f"  {k}. {v}")
        if q["multi"]:
            raw = input("Enter numbers separated by commas: ").strip()
            picks = [p.strip() for p in raw.split(",") if p.strip() in q["options"]]
            profile[q["key"]] = ", ".join(q["options"][p] for p in picks)
        else:
            pick = input("Choice: ").strip()
            profile[q["key"]] = q["options"].get(pick, list(q["options"].values())[0])
    return profile


# ─────────────────────────────────────────────
# GROUND TRUTH LOOKUP
# ─────────────────────────────────────────────

def load_keep_rules(role, profile):
    """Return list of short rule_ids marked KEEP for this role/profile combo."""
    wb = load_workbook(GROUND_TRUTH_XLSX, data_only=True)
    ws = wb["Ground Truth Grid"]
    headers = [c.value for c in ws[1]]
    rule_cols = {h: idx for idx, h in enumerate(headers) if idx >= 4}  # col E onward

    profile_str = "; ".join(f"{k}={v}" for k, v in profile.items())
    target_row = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[1].value == role:
            # Question Dim 1 / Dim 2 columns (C, D) hold the individual profile
            # values, not the full profile_str — match on those instead.
            q1, q2 = row[2].value, row[3].value
            vals = list(profile.values())
            if len(vals) >= 2 and (q1 in vals[0] or vals[0] in str(q1)) and (q2 in vals[1] or vals[1] in str(q2)):
                target_row = row
                break

    if target_row is None:
        print("\n⚠️  Couldn't match this profile to a ground truth row automatically.")
        print("    Falling back to manual row selection.")
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if row[1].value == role:
                print(f"  row {i}: {row[2].value} | {row[3].value}")
        sel = int(input("Enter row number: ").strip())
        target_row = ws[sel]

    keep = []
    for idx, header in enumerate(headers):
        if idx < 4:
            continue
        val = target_row[idx].value
        if val == "KEEP":
            # header is the short display label (e.g. "AIDE E\nBuild DB"),
            # not the rule_id — resolve via Rule Reference sheet.
            keep.append(header)
    return keep, target_row


def resolve_rule_ids(short_headers):
    """Map the Ground Truth Grid's rotated column headers back to full rule IDs
    via the Rule Reference sheet."""
    wb = load_workbook(GROUND_TRUTH_XLSX, data_only=True)
    ws = wb["Rule Reference"]
    # Rule Reference sheet: col A = short key (matches RULE_COLS[i][0] in the
    # build script, NOT the header text) — so instead we rebuild the same
    # RULE_COLS ordering used when the grid was generated.
    RULE_COLS = [
        ("aide_build_database", "AIDE E\nBuild DB"),
        ("aide_periodic_checking_systemd_timer", "AIDE E\nPeriodic"),
        ("partition_for_tmp", "/tmp\nPartition"),
        ("grub2_uefi_password", "GRUB2\nPassword"),
        ("service_systemd-journal-upload_enabled", "journal-\nupload Svc"),
        ("journald_compress", "journalD\nCompress"),
        ("journald_disable_forward_to_syslog", "journalD\nNo-Forward"),
        ("journald_forward_to_syslog", "journalD\nForward"),
        ("journald_storage", "journalD\nStorage"),
        ("socket_systemd-journal-remote_disabled", "journal-\nremote Sock"),
        ("systemd_journal_upload_server_tls", "journal-\nupload TLS"),
        ("systemd_journal_upload_url", "journal-\nupload URL"),
        ("firewall_single_service_active", "Firewall\nSingle Svc"),
        ("service_nftables_enabled", "nftables\nEnabled"),
        ("file_permissions_crontab", "crontab\nPerms"),
        ("package_nis_removed", "NIS\nRemoved"),
        ("package_rpcbind_removed", "rpcbind\nPkg Removed"),
        ("service_rpcbind_disabled", "rpcbind\nSvc Disabled"),
        ("package_ypserv_removed", "ypserv\nPkg Removed"),
        ("service_ypserv_disabled", "ypserv\nSvc Disabled"),
    ]
    header_to_short = {h: s for s, h in RULE_COLS}
    return [PREFIX + header_to_short[h] for h in short_headers if h in header_to_short]


# ─────────────────────────────────────────────
# AGENT-TEST.XML PARSING
# ─────────────────────────────────────────────

def clean_text(elem):
    """Flatten nested XHTML-in-XCCDF text into a plain string."""
    if elem is None:
        return ""
    return re.sub(r"\s+", " ", "".join(elem.itertext())).strip()


def parse_scan_xml(path):
    """
    Returns {rule_id: {"title": str, "description": str, "fix": str, "severity": str}}
    Handles both plain XCCDF results (rule-result only, no embedded Rule/fix —
    in which case fix will be empty) and ARF/results-with-benchmark files
    (which embed the full <Rule> including <fix>). Namespaces are stripped
    for simplicity since XCCDF/ARF namespace URIs vary by SCAP version.
    """
    tree = ET.parse(path)
    root = tree.getroot()

    def strip_ns(tag):
        return tag.split("}")[-1] if "}" in tag else tag

    rules = {}
    for elem in root.iter():
        if strip_ns(elem.tag) != "Rule":
            continue
        rid = elem.get("id", "")
        if not rid:
            continue
        title = None
        desc = None
        fixes = {}  # system -> text, since each Rule can carry sh/ansible/puppet/etc.
        severity = elem.get("severity", "")
        for child in elem:
            ctag = strip_ns(child.tag)
            if ctag == "title" and title is None:
                title = clean_text(child)
            elif ctag == "description" and desc is None:
                desc = clean_text(child)
            elif ctag == "fixtext":
                fixes.setdefault("fixtext", clean_text(child))
            elif ctag == "fix":
                system = child.get("system", "")
                fixes[system] = clean_text(child)

        # Prefer the bash fix (this is what gets bash -c'd), then fixtext,
        # then anything else available (ansible/puppet/blueprint) as reference
        # text only -- some rules genuinely have no automated fix at all.
        fix_text = (
            fixes.get("urn:xccdf:fix:script:sh")
            or fixes.get("fixtext")
            or next(iter(fixes.values()), "")
        )
        fix_system = "sh" if "urn:xccdf:fix:script:sh" in fixes else (
            "fixtext" if "fixtext" in fixes else
            ("other" if fixes else "none")
        )

        rules[rid] = {
            "title": title or rid,
            "description": (desc or "")[:1200],
            "fix": fix_text,
            "fix_system": fix_system,
            "severity": severity,
        }

    if not rules:
        print(f"\n⚠️  No <Rule> elements with description/fix found in {path}.")
        print("    This usually means it's a bare results file (rule-result only,")
        print("    no embedded Benchmark). Re-run oscap with --results-arf instead")
        print("    of --results to get an ARF file with the full rule content, e.g.:")
        print("    sudo oscap xccdf eval --profile <profile> --results-arf agent-test.xml <ds.xml>")
    return rules


# ─────────────────────────────────────────────
# MODEL QUERY
# ─────────────────────────────────────────────

def build_prompt(rule_id, rule_info, role, profile, downtime_sensitivity=None):
    profile_str = "; ".join(f"{k}={v}" for k, v in profile.items())
    live_bash_ok = True
    if downtime_sensitivity and "Ephemeral" in downtime_sensitivity:
        live_bash_ok = False

    style_note = (
        "Give the fix as a config file / Dockerfile-style patch, not live bash commands."
        if not live_bash_ok else
        "Give the fix as a runnable bash script (it will be executed directly with `bash -c`)."
    )

    return f"""You are a Linux system hardening assistant. Produce a remediation script
for the following failed CIS/OpenSCAP rule on this host:

Host: {SYSTEM_INFO['hostname']} | {SYSTEM_INFO['os']} | kernel {SYSTEM_INFO['kernel']} | {SYSTEM_INFO['arch']}
Persona: {role} ({profile_str})

Rule ID: {rule_id}
Title: {rule_info['title']}
Severity: {rule_info['severity']}
Description: {rule_info['description']}
Reference fix from the scan content (format: {rule_info.get('fix_system', 'none')}): {rule_info['fix'] or '(none provided — no automated fix exists for this rule, write the standard remediation yourself)'}

If the reference fix above is bash ("sh"), adapt it directly. If it's Ansible/Puppet/blueprint
or missing, translate the intent into a plain bash script yourself.

{style_note}
Output ONLY the script/config, no prose, no markdown fences, no explanation.
"""


def query_model(model, prompt, max_tokens=800):
    headers = {"Authorization": f"Bearer {NVIDIA_API_KEY}", "Content-Type": "application/json"}
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
        "temperature": 0.2,
    }
    resp = requests.post(NVIDIA_URL, headers=headers, json=payload, timeout=90)
    resp.raise_for_status()
    data = resp.json()
    return data["choices"][0]["message"]["content"].strip()


def strip_code_fences(text):
    text = re.sub(r"^```[a-zA-Z]*\n", "", text.strip())
    text = re.sub(r"\n```$", "", text)
    return text.strip()


# ─────────────────────────────────────────────
# HUMAN CONFIRMATION + EXECUTION
# ─────────────────────────────────────────────

def confirm_and_run(model, rule_id, script):
    print("\n" + "-" * 70)
    print(f"MODEL: {model}")
    print(f"RULE:  {rule_id}")
    print("-" * 70)
    print(script)
    print("-" * 70)
    ans = input("Apply this fix now? [y/n]: ").strip().lower()

    record = {
        "model": model, "rule_id": rule_id, "script": script,
        "approved": ans == "y", "timestamp": datetime.datetime.now().isoformat(),
    }

    if ans != "y":
        record["status"] = "rejected"
        return record

    try:
        proc = subprocess.run(
            ["bash", "-c", script],
            capture_output=True, text=True, timeout=120,
        )
        record["exit_code"] = proc.returncode
        record["stdout"] = proc.stdout[-4000:]
        record["stderr"] = proc.stderr[-4000:]
        record["status"] = "applied_success" if proc.returncode == 0 else "applied_failed"
    except subprocess.TimeoutExpired:
        record["status"] = "applied_timeout"
    except Exception as e:
        record["status"] = "applied_error"
        record["error"] = str(e)

    return record


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    if not NVIDIA_API_KEY:
        print("Set NVIDIA_API_KEY as an environment variable before running:")
        print("  export NVIDIA_API_KEY=nvapi-...")
        sys.exit(1)

    if not os.path.exists(SCAN_RESULT_XML):
        print(f"Missing {SCAN_RESULT_XML} in the current directory.")
        sys.exit(1)
    if not os.path.exists(GROUND_TRUTH_XLSX):
        print(f"Missing {GROUND_TRUTH_XLSX} in the current directory.")
        sys.exit(1)

    role = ask_role()
    profile = ask_followups(role)

    keep_headers, gt_row = load_keep_rules(role, profile)
    keep_rule_ids = resolve_rule_ids(keep_headers)

    print(f"\n{len(keep_rule_ids)} rules marked KEEP for this combo:")
    for rid in keep_rule_ids:
        print(f"  - {rid}")

    scan_rules = parse_scan_xml(SCAN_RESULT_XML)

    os.makedirs(RESULTS_DIR, exist_ok=True)
    run_ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    results = []

    for rule_id in keep_rule_ids:
        rule_info = scan_rules.get(rule_id)
        if rule_info is None:
            print(f"\n⚠️  {rule_id} not found in {SCAN_RESULT_XML} — skipping.")
            continue

        prompt = build_prompt(rule_id, rule_info, role, profile,
                               profile.get("downtime_sensitivity"))

        for model in MODELS:
            print(f"\nQuerying {model} for {rule_id}...")
            try:
                raw = query_model(model, prompt)
            except Exception as e:
                print(f"  request failed: {e}")
                results.append({"model": model, "rule_id": rule_id,
                                 "status": "query_failed", "error": str(e)})
                continue

            script = strip_code_fences(raw)
            record = confirm_and_run(model, rule_id, script)
            results.append(record)
            time.sleep(1)  # be polite to the API

    out_path = os.path.join(RESULTS_DIR, f"remediation_{run_ts}.json")
    json.dump(results, open(out_path, "w"), indent=2)

    print("\n" + "=" * 55)
    print("  SCOREBOARD")
    print("=" * 55)
    for model in MODELS:
        model_rows = [r for r in results if r["model"] == model]
        approved = sum(1 for r in model_rows if r.get("approved"))
        success = sum(1 for r in model_rows if r.get("status") == "applied_success")
        failed = sum(1 for r in model_rows if r.get("status") == "applied_failed")
        rejected = sum(1 for r in model_rows if r.get("status") == "rejected")
        print(f"{model}")
        print(f"  attempted: {len(model_rows)} | approved: {approved} | "
              f"succeeded: {success} | failed: {failed} | rejected: {rejected}")

    print(f"\nFull log: {out_path}")


if __name__ == "__main__":
    main()