#!/usr/bin/env python3
"""
CIS Benchmark Multi-Model Remediation Pipeline v3.0
=====================================================
Key changes from v2:
  - Uses LAB SERVER (Ollama over VPN) instead of NVIDIA NIM
  - Sequential per-model mode: ALL rules for Model A -> snapshot restore
    prompt -> ALL rules for Model B -> etc.
  - oscap single-rule verification after each fix attempt
  - Records PASS/FAIL per rule per model in a comparison table
  - Auto-generates break_rules.sh to reset state between models
  - Full results saved as JSON + comparison.md

Lab models (from your benchmarking data):
  qwen2.5:7b   gemma2:latest   mistral:latest
  granite4.1:8b   gpt-oss:latest

Usage:
  python3 remediationv2.py                       # interactive, human approval
  python3 remediationv2.py --auto                # auto-approve all fixes
  python3 remediationv2.py --model qwen2.5:7b   # single model only
  python3 remediationv2.py --snapshot            # prompt snapshot restore between models
"""

import os, re, sys, json, time, datetime, subprocess, xml.etree.ElementTree as ET
import requests, urllib3
from openpyxl import load_workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

LAB_URL  = os.environ.get("LAB_URL",  "https://10.1.96.96:8443")
LAB_USER = os.environ.get("LAB_USER", "user")
LAB_PASS = os.environ.get("LAB_PASS", "H72j8n19sna")

MODELS = [
    "qwen2.5:7b",
    "gemma2:latest",
    "mistral:latest",
    "granite4.1:8b",
    "gpt-oss:latest",
]

SCAN_RESULT_XML   = "agent-test.xml"
GROUND_TRUTH_XLSX = "CIS_Ground_Truth_FINAL.xlsx"
RESULTS_DIR       = "remediation_results"
BENCHMARK_XML     = os.path.expanduser(
    "~/Downloads/scap-security-guide-0.1.76/ssg-ubuntu2404-ds.xml")
SNAPSHOT_NAME     = "baseline-63-rules-broken"

SYSTEM_INFO = {
    "hostname": "pranjal-garg-VirtualBox",
    "kernel":   "6.17.0-23-generic",
    "os":       "Ubuntu 24.04 LTS (Noble Numbat)",
    "arch":     "x86_64",
}

PREFIX = "xccdf_org.ssgproject.content_rule_"

# ─────────────────────────────────────────────────────────────────────────────
# RULE COLS — must match your Ground Truth Grid column order exactly
# ─────────────────────────────────────────────────────────────────────────────

RULE_COLS = [
    # Original 20 rules
    ("aide_build_database",                        "AIDE E\nBuild DB"),
    ("aide_periodic_checking_systemd_timer",       "AIDE E\nPeriodic"),
    ("partition_for_tmp",                          "/tmp\nPartition"),
    ("grub2_uefi_password",                        "GRUB2\nPassword"),
    ("service_systemd-journal-upload_enabled",     "journal-\nupload Svc"),
    ("journald_compress",                          "journalD\nCompress"),
    ("journald_disable_forward_to_syslog",         "journalD\nNo-Forward"),
    ("journald_forward_to_syslog",                 "journalD\nForward"),
    ("journald_storage",                           "journalD\nStorage"),
    ("socket_systemd-journal-remote_disabled",     "journal-\nremote Sock"),
    ("systemd_journal_upload_server_tls",          "journal-\nupload TLS"),
    ("systemd_journal_upload_url",                 "journal-\nupload URL"),
    ("firewall_single_service_active",             "Firewall\nSingle Svc"),
    ("service_nftables_enabled",                   "nftables\nEnabled"),
    ("file_permissions_crontab",                   "crontab\nPerms"),
    ("package_nis_removed",                        "NIS\nRemoved"),
    ("package_rpcbind_removed",                    "rpcbind\nPkg Removed"),
    ("service_rpcbind_disabled",                   "rpcbind\nSvc Disabled"),
    ("package_ypserv_removed",                     "ypserv\nPkg Removed"),
    ("service_ypserv_disabled",                    "ypserv\nSvc Disabled"),
    # New 40 rules
    ("accounts_passwords_pam_faillock_deny",       "PAM\nfaillock deny"),
    ("accounts_passwords_pam_faillock_enabled",    "PAM\nfaillock ena"),
    ("accounts_passwords_pam_faillock_unlock_time","PAM\nunlock time"),
    ("accounts_password_pam_dcredit",              "PAM\ndcredit"),
    ("accounts_password_pam_minlen",               "PAM\nminlen"),
    ("accounts_password_pam_ucredit",              "PAM\nucredit"),
    ("accounts_password_pam_unix_no_remember",     "PAM\nno_remember"),
    ("no_empty_passwords_unix",                    "No Empty\nPasswords"),
    ("accounts_tmout",                             "Session\nTimeout"),
    ("accounts_umask_etc_bashrc",                  "Umask\nbashrc"),
    ("sudo_custom_logfile",                        "sudo\nLogfile"),
    ("sudo_require_reauthentication",              "sudo\nReauth"),
    ("package_apparmor-utils_installed",           "AppArmor\nUtils"),
    ("grub2_enable_apparmor",                      "AppArmor\nGRUB"),
    ("sysctl_net_ipv6_conf_all_forwarding",        "IPv6\nForwarding"),
    ("sysctl_net_ipv4_conf_all_accept_redirects",  "IPv4\nAccept Redir"),
    ("sysctl_net_ipv4_conf_all_log_martians",      "IPv4\nLog Martians"),
    ("sysctl_net_ipv4_conf_all_rp_filter",         "IPv4\nRP Filter"),
    ("sysctl_net_ipv4_tcp_syncookies",             "IPv4\nSyncookies"),
    ("sysctl_net_ipv4_conf_all_send_redirects",    "IPv4\nSend Redir"),
    ("sysctl_net_ipv4_ip_forward",                 "IPv4\nIP Forward"),
    ("sysctl_kernel_randomize_va_space",           "Kernel\nASLR"),
    ("sysctl_fs_suid_dumpable",                    "suid\nDumpable"),
    ("disable_users_coredumps",                    "Core\nDumps"),
    ("kernel_module_cramfs_disabled",              "cramfs\nDisabled"),
    ("kernel_module_hfs_disabled",                 "hfs\nDisabled"),
    ("kernel_module_jffs2_disabled",               "jffs2\nDisabled"),
    ("mount_option_dev_shm_nodev",                 "/dev/shm\nnodev"),
    ("mount_option_dev_shm_noexec",                "/dev/shm\nnoexec"),
    ("mount_option_dev_shm_nosuid",                "/dev/shm\nnosuid"),
    ("file_permissions_cron_allow",                "cron.allow\nPerms"),
    ("file_permissions_cron_d",                    "cron.d\nPerms"),
    ("file_permissions_cron_daily",                "cron.daily\nPerms"),
    ("file_owner_cron_allow",                      "cron.allow\nOwner"),
    ("package_ftp_removed",                        "ftp\nRemoved"),
    ("package_openldap-clients_removed",           "LDAP\nRemoved"),
    ("package_rsync_removed",                      "rsync\nRemoved"),
    ("service_rsyncd_disabled",                    "rsyncd\nDisabled"),
    ("package_telnet_removed",                     "telnet\nRemoved"),
    ("all_apparmor_profiles_in_enforce_complain_mode", "AppArmor\nEnforce"),
]

# ─────────────────────────────────────────────────────────────────────────────
# ROLE / FOLLOW-UP QUESTIONS
# ─────────────────────────────────────────────────────────────────────────────

ROLES = {
    "1": "Personal Laptop / Home User",
    "2": "Student / Security Learner / Researcher",
    "3": "Software Developer",
    "4": "System / Cloud Administrator",
}

FOLLOWUP_QUESTIONS = {
    "Personal Laptop / Home User": [
        {"q": "Who physically uses this computer?",
         "options": {
             "1": "Just me (Low risk of physical tampering)",
             "2": "Shared with family or roommates (Moderate risk, needs basic user isolation)"},
         "key": "physical_access", "multi": False},
        {"q": "Where do you connect?",
         "options": {
             "1": "Only trusted home/private networks (Standard firewall is fine)",
             "2": "Frequently on public campus or cafe Wi-Fi (Needs aggressive network hardening)"},
         "key": "network_environment", "multi": False},
    ],
    "Student / Security Learner / Researcher": [
        {"q": "What do you actually use this computer for? (Select ALL that apply)",
         "options": {
             "1": "Coding & Development (Writing code, running local web servers, or building apps)",
             "2": "Security & Hacking (Playing with network scanners, testing vulnerabilities, or CTFs)",
             "3": "General Technical Work (Basic scripting, data analysis, and standard terminal usage)"},
         "key": "learning_workloads", "multi": True},
        {"q": "How comfortable are you with the Linux terminal?",
         "options": {
             "1": "Beginner (Explain exactly what the commands do before I run them)",
             "2": "Advanced (Just give me the raw commands or config file edits, I know what they do)"},
         "key": "technical_depth", "multi": False},
    ],
    "Software Developer": [
        {"q": "What are you building? (Select ALL that apply)",
         "options": {
             "1": "Web / Full-Stack (MERN, React Native, Node.js - needs local port access)",
             "2": "Systems / Low-Level (C/C++, Linux kernels - needs deep system execution rights)",
             "3": "Containerized Apps (Docker/Podman - relies on virtual networking)"},
         "key": "dev_stack", "multi": True},
        {"q": "Does this machine accept external connections?",
         "options": {
             "1": "Yes, I run local servers/APIs that teammates or external tools connect to",
             "2": "No, strictly offline compiling and local-only testing"},
         "key": "network_exposure", "multi": False},
    ],
    "System / Cloud Administrator": [
        {"q": "How sensitive is this server to downtime?",
         "options": {
             "1": "Production / Critical",
             "2": "Internal / Workstation",
             "3": "Ephemeral (config/Dockerfile fixes only, no live bash)"},
         "key": "downtime_sensitivity", "multi": False},
        {"q": "Where does this infrastructure live?",
         "options": {
             "1": "Public Cloud",
             "2": "Internal Corporate Network",
             "3": "Local Virtual Machine (Sandboxed environment)"},
         "key": "infrastructure_location", "multi": False},
    ],
}


def ask_role():
    print("\n" + "=" * 60)
    print("  CIS Remediation Pipeline v3.0 — select your role")
    print("=" * 60)
    for k, v in ROLES.items():
        print(f"  {k}. {v}")
    while True:
        c = input("\nRole number: ").strip()
        if c in ROLES:
            return ROLES[c]
        print("Invalid, try again.")


def ask_followups(role):
    questions = FOLLOWUP_QUESTIONS.get(role, [])
    profile = {}
    for q in questions:
        print(f"\n{q['q']}")
        for k, v in q["options"].items():
            print(f"  {k}. {v}")
        if q["multi"]:
            raw = input("Enter numbers separated by commas: ").strip()
            picks = [p.strip() for p in raw.split(",") if p.strip() in q["options"]]
            profile[q["key"]] = ", ".join(q["options"][p] for p in picks)
        else:
            pick = input("Choice: ").strip()
            profile[q["key"]] = q["options"].get(pick, list(q["options"].values())[0])
    return profile


# ─────────────────────────────────────────────────────────────────────────────
# GROUND TRUTH LOOKUP
# ─────────────────────────────────────────────────────────────────────────────

def load_keep_rules(role, profile):
    wb = load_workbook(GROUND_TRUTH_XLSX, data_only=True)
    ws = wb["Ground Truth Grid"]

    target_row = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[1].value != role:
            continue
        q1  = str(row[2].value or "")
        q2  = str(row[3].value or "")
        vals = list(profile.values())
        v0  = str(vals[0]) if len(vals) > 0 else ""
        v1  = str(vals[1]) if len(vals) > 1 else ""
        if (q1 in v0 or v0 in q1) and (q2 in v1 or v1 in q2):
            target_row = row
            break

    if target_row is None:
        print("\n  Could not auto-match profile. Manual row selection:")
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
            if row[1].value == role:
                print(f"  row {i}: {row[2].value} | {row[3].value}")
        sel = int(input("Enter row number: ").strip())
        target_row = list(ws.iter_rows(min_row=sel, max_row=sel))[0]

    keep_ids = []
    for idx, (short_id, _) in enumerate(RULE_COLS):
        col_idx = idx + 4
        if col_idx >= len(target_row):
            continue
        if target_row[col_idx].value == "KEEP":
            keep_ids.append(PREFIX + short_id)
    return keep_ids


# ─────────────────────────────────────────────────────────────────────────────
# SCAN XML PARSING
# ─────────────────────────────────────────────────────────────────────────────

def parse_scan_xml(path):
    rules = {}
    try:
        tree = ET.parse(path)
    except Exception as e:
        print(f"[ERROR] Cannot parse {path}: {e}")
        return rules

    root = tree.getroot()
    for ns_uri in ["http://checklists.nist.gov/xccdf/1.2",
                   "http://checklists.nist.gov/xccdf/1.1"]:
        for rule_el in root.iter(f"{{{ns_uri}}}Rule"):
            rid = rule_el.get("id", "")
            if not rid.startswith(PREFIX):
                continue

            title_el = rule_el.find(f"{{{ns_uri}}}title")
            title    = (title_el.text or "").strip() if title_el is not None else rid

            desc_el  = rule_el.find(f"{{{ns_uri}}}description")
            desc     = "".join(desc_el.itertext()).strip() if desc_el is not None else ""

            severity = rule_el.get("severity", "unknown")

            fixes = {}
            for fix_el in rule_el.findall(f"{{{ns_uri}}}fix"):
                fixes[fix_el.get("system", "fixtext")] = (fix_el.text or "").strip()
            ft = rule_el.find(f"{{{ns_uri}}}fixtext")
            if ft is not None and ft.text:
                fixes["fixtext"] = ft.text.strip()

            fix_text = (
                fixes.get("urn:xccdf:fix:script:sh")
                or fixes.get("fixtext")
                or next(iter(fixes.values()), "")
            )
            fix_system = (
                "sh"      if "urn:xccdf:fix:script:sh" in fixes else
                "fixtext" if "fixtext" in fixes else
                ("other"  if fixes else "none")
            )

            rules[rid] = {
                "title":       title,
                "description": desc[:1200],
                "fix":         fix_text,
                "fix_system":  fix_system,
                "severity":    severity,
            }

    if not rules:
        print(f"\n  No Rule elements found in {path}.")
        print("  Re-run oscap with --results-arf to embed full benchmark content.")
    return rules


# ─────────────────────────────────────────────────────────────────────────────
# LAB SERVER QUERY
# ─────────────────────────────────────────────────────────────────────────────

def query_lab(model, prompt, max_tokens=900):
    payload = {
        "model":       model,
        "messages":    [{"role": "user", "content": prompt}],
        "max_tokens":  max_tokens,
        "temperature": 0.2,
        "stream":      False,
    }
    resp = requests.post(
        f"{LAB_URL}/v1/chat/completions",
        json=payload,
        auth=(LAB_USER, LAB_PASS),
        verify=False,
        timeout=120,
    )
    resp.raise_for_status()
    data = resp.json()
    if "choices" in data:
        return data["choices"][0]["message"]["content"].strip()
    elif "message" in data:
        return data["message"]["content"].strip()
    else:
        raise ValueError(f"Unexpected response: {list(data.keys())}")


def strip_code_fences(text):
    text = re.sub(r"^```[a-zA-Z]*\n?", "", text.strip())
    text = re.sub(r"\n?```$", "", text)
    return text.strip()


# ─────────────────────────────────────────────────────────────────────────────
# PROMPT BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_prompt(rule_id, rule_info, role, profile):
    profile_str = "; ".join(f"{k}={v}" for k, v in profile.items())
    ephemeral   = "Ephemeral" in profile.get("downtime_sensitivity", "")
    style_note  = (
        "Give the fix as a config file / Dockerfile-style patch — NOT live bash."
        if ephemeral else
        "Give the fix as a runnable bash script (it will be run directly with `bash -c`)."
    )
    return f"""You are a Linux system hardening expert. Write a remediation script for
the following failed CIS/OpenSCAP rule on this specific host:

Host: {SYSTEM_INFO['hostname']} | {SYSTEM_INFO['os']} | kernel {SYSTEM_INFO['kernel']} | {SYSTEM_INFO['arch']}
Persona: {role} ({profile_str})

Rule ID:     {rule_id}
Title:       {rule_info['title']}
Severity:    {rule_info['severity']}
Description: {rule_info['description']}

Reference fix from benchmark (format: {rule_info.get('fix_system','none')}):
{rule_info['fix'] or '(none provided — write the standard Ubuntu 24.04 remediation yourself)'}

If the reference fix is bash (sh), adapt it directly.
If it is Ansible/Puppet/blueprint or missing, translate the intent into plain bash.

{style_note}
Output ONLY the script — no prose, no markdown fences, no explanation.
"""


# ─────────────────────────────────────────────────────────────────────────────
# OSCAP SINGLE-RULE VERIFICATION
# ─────────────────────────────────────────────────────────────────────────────

def verify_rule(rule_id):
    cmd = (
        f"sudo oscap xccdf eval --rule {rule_id} "
        f"{BENCHMARK_XML} 2>/dev/null | grep 'Result'"
    )
    try:
        out = subprocess.check_output(
            cmd, shell=True, text=True, timeout=90
        ).strip().lower()
        if "pass" in out:            return "pass"
        elif "fail" in out:          return "fail"
        elif "notapplicable" in out: return "notapplicable"
        else:                        return "notchecked"
    except subprocess.TimeoutExpired:
        return "timeout"
    except Exception:
        return "error"


# ─────────────────────────────────────────────────────────────────────────────
# BREAK-RULES SCRIPT GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

BREAK_COMMANDS = {
    "accounts_passwords_pam_faillock_deny":
        "sudo sed -i '/faillock/d' /etc/pam.d/common-auth /etc/pam.d/common-account 2>/dev/null || true",
    "accounts_passwords_pam_faillock_enabled":
        "sudo sed -i '/pam_faillock/d' /etc/pam.d/common-auth /etc/pam.d/common-account 2>/dev/null || true",
    "accounts_passwords_pam_faillock_unlock_time":
        "sudo sed -i '/faillock/d' /etc/pam.d/common-auth 2>/dev/null || true",
    "accounts_password_pam_dcredit":
        "sudo sed -i '/dcredit/d' /etc/security/pwquality.conf 2>/dev/null || true",
    "accounts_password_pam_minlen":
        "sudo sed -i '/minlen/d' /etc/security/pwquality.conf 2>/dev/null || true",
    "accounts_password_pam_ucredit":
        "sudo sed -i '/ucredit/d' /etc/security/pwquality.conf 2>/dev/null || true",
    "accounts_password_pam_unix_no_remember":
        "sudo sed -i '/remember/d' /etc/pam.d/common-password 2>/dev/null || true",
    "no_empty_passwords_unix":
        "sudo sed -i 's/ nullok_secure//g' /etc/pam.d/common-auth 2>/dev/null || true",
    "accounts_tmout":
        "sudo sed -i '/TMOUT/d' /etc/bash.bashrc /etc/profile /etc/profile.d/*.sh 2>/dev/null || true",
    "accounts_umask_etc_bashrc":
        "sudo sed -i '/umask 027/d; /umask 077/d' /etc/bash.bashrc 2>/dev/null || true",
    "sudo_custom_logfile":
        "sudo sed -i '/logfile/d' /etc/sudoers 2>/dev/null; sudo rm -f /etc/sudoers.d/logfile 2>/dev/null || true",
    "sudo_require_reauthentication":
        "sudo rm -f /etc/sudoers.d/timeout 2>/dev/null; echo 'Defaults timestamp_timeout=15' | sudo tee /etc/sudoers.d/noauth > /dev/null",
    "package_apparmor-utils_installed":
        "sudo apt-get remove -y apparmor-utils 2>/dev/null || true",
    "grub2_enable_apparmor":
        "sudo sed -i 's/ apparmor=1 security=apparmor//' /etc/default/grub 2>/dev/null; sudo update-grub 2>/dev/null || true",
    "all_apparmor_profiles_in_enforce_complain_mode":
        "sudo aa-complain /etc/apparmor.d/* 2>/dev/null || true",
    "sysctl_net_ipv6_conf_all_forwarding":
        "sudo sed -i '/net.ipv6.conf.all.forwarding/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv6.conf.all.forwarding=1 2>/dev/null || true",
    "sysctl_net_ipv4_conf_all_accept_redirects":
        "sudo sed -i '/accept_redirects/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.conf.all.accept_redirects=1 2>/dev/null || true",
    "sysctl_net_ipv4_conf_all_log_martians":
        "sudo sed -i '/log_martians/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.conf.all.log_martians=0 2>/dev/null || true",
    "sysctl_net_ipv4_conf_all_rp_filter":
        "sudo sed -i '/rp_filter/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.conf.all.rp_filter=0 2>/dev/null || true",
    "sysctl_net_ipv4_tcp_syncookies":
        "sudo sed -i '/syncookies/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.tcp_syncookies=0 2>/dev/null || true",
    "sysctl_net_ipv4_conf_all_send_redirects":
        "sudo sed -i '/send_redirects/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.conf.all.send_redirects=1 2>/dev/null || true",
    "sysctl_net_ipv4_ip_forward":
        "sudo sed -i '/ip_forward/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w net.ipv4.ip_forward=1 2>/dev/null || true",
    "sysctl_kernel_randomize_va_space":
        "sudo sed -i '/randomize_va_space/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w kernel.randomize_va_space=0 2>/dev/null || true",
    "sysctl_fs_suid_dumpable":
        "sudo sed -i '/suid_dumpable/d' /etc/sysctl.conf /etc/sysctl.d/*.conf 2>/dev/null; sudo sysctl -w fs.suid_dumpable=1 2>/dev/null || true",
    "disable_users_coredumps":
        "sudo sed -i '/hard core/d; /soft core/d' /etc/security/limits.conf 2>/dev/null; echo '* soft core unlimited' | sudo tee -a /etc/security/limits.conf",
    "kernel_module_cramfs_disabled":
        "sudo rm -f /etc/modprobe.d/cramfs.conf 2>/dev/null || true",
    "kernel_module_hfs_disabled":
        "sudo rm -f /etc/modprobe.d/hfs.conf 2>/dev/null || true",
    "kernel_module_jffs2_disabled":
        "sudo rm -f /etc/modprobe.d/jffs2.conf 2>/dev/null || true",
    "mount_option_dev_shm_nodev":
        "sudo sed -i '/\\/dev\\/shm/d' /etc/fstab; sudo mount -o remount /dev/shm 2>/dev/null || true",
    "mount_option_dev_shm_noexec":
        "sudo sed -i '/\\/dev\\/shm/d' /etc/fstab; sudo mount -o remount /dev/shm 2>/dev/null || true",
    "mount_option_dev_shm_nosuid":
        "sudo sed -i '/\\/dev\\/shm/d' /etc/fstab; sudo mount -o remount /dev/shm 2>/dev/null || true",
    "file_permissions_crontab":
        "sudo chmod 0644 /etc/crontab 2>/dev/null || true",
    "file_permissions_cron_allow":
        "sudo chmod 0644 /etc/cron.allow 2>/dev/null || true",
    "file_permissions_cron_d":
        "sudo chmod 0755 /etc/cron.d 2>/dev/null || true",
    "file_permissions_cron_daily":
        "sudo chmod 0755 /etc/cron.daily 2>/dev/null || true",
    "file_owner_cron_allow":
        "sudo chown nobody:nogroup /etc/cron.allow 2>/dev/null || true",
    "package_ftp_removed":
        "sudo apt-get install -y ftp 2>/dev/null || true",
    "package_openldap-clients_removed":
        "sudo apt-get install -y ldap-utils 2>/dev/null || true",
    "package_rsync_removed":
        "sudo apt-get install -y rsync 2>/dev/null || true",
    "service_rsyncd_disabled":
        "sudo systemctl enable rsync 2>/dev/null || sudo systemctl enable rsyncd 2>/dev/null || true",
    "package_telnet_removed":
        "sudo apt-get install -y telnet 2>/dev/null || true",
}


def generate_break_script(keep_rule_ids, output_path="break_rules.sh"):
    not_found = []
    lines = [
        "#!/bin/bash",
        "# Auto-generated by remediationv2.py v3.0",
        "# Resets all benchmark rules back to FAILING state",
        "# Run between models: bash break_rules.sh",
        "echo '=== Resetting rules to failing state ==='",
        "",
    ]
    for rule_id in keep_rule_ids:
        short = rule_id.replace(PREFIX, "")
        if short in BREAK_COMMANDS:
            lines.append(f"echo '  Breaking: {short}'")
            lines.append(BREAK_COMMANDS[short])
        else:
            not_found.append(short)

    lines += [
        "",
        "echo '=== Reset complete ==='",
    ]
    if not_found:
        lines.append(f"echo 'No break command for: {not_found}'")

    with open(output_path, "w") as f:
        f.write("\n".join(lines) + "\n")
    os.chmod(output_path, 0o755)

    print(f"\n  Break script written to: {output_path}")
    if not_found:
        print(f"  Rules without break command (add manually): {not_found}")
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# SINGLE MODEL RUN
# ─────────────────────────────────────────────────────────────────────────────

def run_one_model(model, keep_rule_ids, scan_rules, role, profile, auto_approve):
    print(f"\n{'='*60}")
    print(f"  MODEL : {model}")
    print(f"  Rules : {len(keep_rule_ids)}")
    print(f"{'='*60}")

    model_results = {
        "model":   model,
        "role":    role,
        "profile": profile,
        "started": datetime.datetime.now().isoformat(),
        "rules":   {},
        "summary": {
            "attempted": 0, "approved": 0,
            "script_ok": 0, "oscap_pass": 0, "oscap_fail": 0,
            "query_error": 0, "script_error": 0, "rejected": 0,
        },
    }

    for rule_id in keep_rule_ids:
        short     = rule_id.replace(PREFIX, "")
        rule_info = scan_rules.get(rule_id)

        print(f"\n  [{short}]")
        if rule_info is None:
            print("    Not found in scan XML — skipping")
            model_results["rules"][rule_id] = {"status": "not_in_scan"}
            continue

        # Query lab model
        print(f"    Querying {model}...")
        try:
            raw = query_lab(model, build_prompt(rule_id, rule_info, role, profile))
        except Exception as e:
            print(f"    [QUERY ERROR] {e}")
            model_results["rules"][rule_id] = {"status": "query_error", "error": str(e)}
            model_results["summary"]["query_error"] += 1
            time.sleep(2)
            continue

        script = strip_code_fences(raw)
        model_results["summary"]["attempted"] += 1

        # Show script
        print("\n    Proposed fix:")
        print("    " + "-"*52)
        for line in script.split("\n")[:20]:
            print(f"    {line}")
        if script.count("\n") > 20:
            print(f"    ... ({script.count(chr(10))-20} more lines)")
        print("    " + "-"*52)

        # Approval
        if auto_approve:
            approved = True
            print("    [AUTO] Applying.")
        else:
            ans = input("    Apply? [y/n/s=show full]: ").strip().lower()
            if ans == "s":
                print(f"\n{script}\n")
                ans = input("    Apply? [y/n]: ").strip().lower()
            approved = (ans == "y")

        rule_record = {"script": script, "approved": approved}

        if not approved:
            rule_record["status"] = "rejected"
            model_results["summary"]["rejected"] += 1
            model_results["rules"][rule_id] = rule_record
            continue

        model_results["summary"]["approved"] += 1

        # Apply fix
        try:
            proc = subprocess.run(
                ["bash", "-c", script],
                capture_output=True, text=True, timeout=120,
            )
            rule_record["exit_code"] = proc.returncode
            rule_record["stdout"]    = proc.stdout[-2000:]
            rule_record["stderr"]    = proc.stderr[-2000:]

            if proc.returncode != 0:
                print(f"    [SCRIPT ERROR] exit={proc.returncode}")
                print(f"    {proc.stderr[-150:]}")
                rule_record["status"] = "script_error"
                model_results["summary"]["script_error"] += 1
            else:
                model_results["summary"]["script_ok"] += 1
                print("    Verifying with oscap...")
                verdict = verify_rule(rule_id)
                rule_record["oscap_result"] = verdict
                if verdict == "pass":
                    print("    PASS")
                    rule_record["status"] = "oscap_pass"
                    model_results["summary"]["oscap_pass"] += 1
                else:
                    print(f"    {verdict.upper()}")
                    rule_record["status"] = f"oscap_{verdict}"
                    model_results["summary"]["oscap_fail"] += 1

        except subprocess.TimeoutExpired:
            print("    [TIMEOUT]")
            rule_record["status"] = "timeout"
            model_results["summary"]["script_error"] += 1
        except Exception as e:
            rule_record["status"] = "error"
            rule_record["error"]  = str(e)
            model_results["summary"]["script_error"] += 1

        model_results["rules"][rule_id] = rule_record
        time.sleep(1)

    # Print model summary
    s     = model_results["summary"]
    total = len(keep_rule_ids)
    pct   = (s["oscap_pass"] / total * 100) if total else 0
    print(f"\n  {model} DONE")
    print(f"  Attempted={s['attempted']} Approved={s['approved']} "
          f"PASS={s['oscap_pass']}({pct:.1f}%) FAIL={s['oscap_fail']} "
          f"Err={s['script_error']+s['query_error']}")

    model_results["finished"] = datetime.datetime.now().isoformat()
    return model_results


# ─────────────────────────────────────────────────────────────────────────────
# RESET PROMPT BETWEEN MODELS
# ─────────────────────────────────────────────────────────────────────────────

def prompt_reset(prev_model, next_model, break_script, use_snapshot):
    print(f"\n{'='*60}")
    print(f"  {prev_model} complete. Next: {next_model}")
    print(f"{'='*60}")
    if use_snapshot:
        print(f"""
  RESTORE VM SNAPSHOT before continuing.
  Snapshot name: '{SNAPSHOT_NAME}'

  On your HOST machine:
    VBoxManage controlvm "pranjal-garg-VirtualBox" poweroff
    VBoxManage snapshot "pranjal-garg-VirtualBox" restore "{SNAPSHOT_NAME}"
    VBoxManage startvm "pranjal-garg-VirtualBox" --type headless

  Then SSH back in and re-run with --model "{next_model}" --auto
""")
    else:
        print(f"""
  RESET RULES TO FAILING STATE before continuing.
  Run on the VM:
    bash {break_script}

  Then verify rules are failing:
    sudo oscap xccdf eval \\
      --profile xccdf_org.ssgproject.content_profile_cis_level1_workstation \\
      --results agent-test.xml \\
      ~/Downloads/scap-security-guide-0.1.76/ssg-ubuntu2404-ds.xml
""")
    input("  Press Enter when ready for next model...")


# ─────────────────────────────────────────────────────────────────────────────
# SAVE RESULTS
# ─────────────────────────────────────────────────────────────────────────────

def save_results(all_results, role, profile):
    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    slug     = role.lower().replace(" ", "_").replace("/", "_")
    run_dir  = os.path.join(RESULTS_DIR, f"remediation_{slug}_{ts}")
    os.makedirs(run_dir, exist_ok=True)

    for r in all_results:
        fname = r["model"].replace(":", "_").replace("/", "_") + ".json"
        with open(os.path.join(run_dir, fname), "w") as f:
            json.dump(r, f, indent=2)

    # Comparison markdown
    lines = [
        "# CIS Remediation Multi-Model Comparison\n\n",
        f"**Role:** {role}\n\n",
        "**Profile:**\n",
    ]
    for k, v in profile.items():
        lines.append(f"- {k}: {v}\n")
    lines.append(f"\n**Timestamp:** {ts}\n\n---\n\n")

    # Scoreboard
    lines.append("## Scoreboard\n\n")
    lines.append("| Model | Rules | PASS | PASS% | FAIL | Errors |\n")
    lines.append("|---|---|---|---|---|---|\n")
    for r in all_results:
        s   = r["summary"]
        tot = len(r["rules"])
        pct = (s["oscap_pass"]/tot*100) if tot else 0
        lines.append(f"| {r['model']} | {tot} | {s['oscap_pass']} | "
                     f"{pct:.1f}% | {s['oscap_fail']} | "
                     f"{s['script_error']+s['query_error']} |\n")

    lines.append("\n---\n\n## Per-Rule Results\n\n")

    all_rids = []
    for r in all_results:
        for rid in r["rules"]:
            if rid not in all_rids:
                all_rids.append(rid)

    hdr = "| Rule | " + " | ".join(r["model"] for r in all_results) + " |\n"
    sep = "|---|" + "---|" * len(all_results) + "\n"
    lines += [hdr, sep]

    for rid in all_rids:
        short = rid.replace(PREFIX, "")
        row   = f"| {short} |"
        for r in all_results:
            rec    = r["rules"].get(rid, {})
            status = rec.get("status", "-")
            cell   = (
                "PASS" if status == "oscap_pass" else
                "FAIL" if "oscap_fail" in status else
                status.replace("_", " ")
            )
            row += f" {cell} |"
        lines.append(row + "\n")

    with open(os.path.join(run_dir, "comparison.md"), "w") as f:
        f.writelines(lines)

    with open(os.path.join(run_dir, "summary.json"), "w") as f:
        json.dump({
            "role": role, "profile": profile, "timestamp": ts,
            "models": [{"model": r["model"], **r["summary"]} for r in all_results],
        }, f, indent=2)

    return run_dir


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    auto_approve = "--auto"     in sys.argv
    use_snapshot = "--snapshot" in sys.argv
    single_model = None

    if "--model" in sys.argv:
        idx = sys.argv.index("--model")
        if idx + 1 < len(sys.argv):
            single_model = sys.argv[idx + 1]

    print("\n" + "=" * 60)
    print("  CIS Multi-Model Remediation Pipeline v3.0")
    print(f"  Lab server : {LAB_URL}")
    print(f"  Models     : {', '.join(MODELS)}")
    print("=" * 60)

    if auto_approve:  print("  [AUTO]     All scripts applied without confirmation.")
    if single_model:  print(f"  [SINGLE]   Running only: {single_model}")
    if use_snapshot:  print(f"  [SNAPSHOT] Will prompt snapshot restore between models.")
    else:             print("  [BREAK]    Will generate break_rules.sh between models.")

    for f in [SCAN_RESULT_XML, GROUND_TRUTH_XLSX]:
        if not os.path.exists(f):
            print(f"\n[ERROR] Missing: {f}")
            sys.exit(1)

    role    = ask_role()
    profile = ask_followups(role)

    keep_rule_ids = load_keep_rules(role, profile)
    print(f"\n  {len(keep_rule_ids)} KEEP rules for this combo:")
    for rid in keep_rule_ids:
        print(f"    {rid.replace(PREFIX,'')}")

    scan_rules   = parse_scan_xml(SCAN_RESULT_XML)
    break_script = generate_break_script(keep_rule_ids)

    models_to_run = MODELS if not single_model else [
        m for m in MODELS if m == single_model
    ]
    if not models_to_run:
        print(f"[ERROR] '{single_model}' not in MODELS list.")
        sys.exit(1)

    os.makedirs(RESULTS_DIR, exist_ok=True)
    all_results = []

    print(f"\n  Running {len(models_to_run)} model(s) sequentially.")
    input("  Press Enter to start with the first model...")

    for i, model in enumerate(models_to_run):
        if i > 0:
            prompt_reset(
                prev_model   = models_to_run[i-1],
                next_model   = model,
                break_script = break_script,
                use_snapshot = use_snapshot,
            )

        result = run_one_model(
            model         = model,
            keep_rule_ids = keep_rule_ids,
            scan_rules    = scan_rules,
            role          = role,
            profile       = profile,
            auto_approve  = auto_approve,
        )
        all_results.append(result)

        # Save after each model so a crash doesn't lose data
        ts    = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = os.path.join(
            RESULTS_DIR,
            f"{model.replace(':','_').replace('/','_')}_{ts}.json"
        )
        with open(fname, "w") as f:
            json.dump(result, f, indent=2)
        print(f"  Saved: {fname}")

    run_dir = save_results(all_results, role, profile)

    # Final scoreboard
    print(f"\n{'='*60}")
    print("  FINAL SCOREBOARD")
    print(f"{'='*60}")
    print(f"  {'Model':<25} {'Rules':>6} {'PASS':>6} {'PASS%':>7} {'ERR':>5}")
    print(f"  {'-'*25} {'----':>6} {'----':>6} {'-----':>7} {'---':>5}")
    for r in all_results:
        s   = r["summary"]
        tot = len(r["rules"])
        pct = (s["oscap_pass"]/tot*100) if tot else 0
        print(f"  {r['model']:<25} {tot:>6} {s['oscap_pass']:>6} "
              f"{pct:>6.1f}% {s['script_error']+s['query_error']:>5}")

    print(f"\n  Results : {run_dir}")
    print(f"  Table   : {run_dir}/comparison.md")
    print(f"  Reset   : bash {break_script}")


if __name__ == "__main__":
    main()
